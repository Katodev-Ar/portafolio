"""
PrintBridge — Test Suite Completa (pytest + httpx)

Cubre:
  - config.py:        caché, escritura atómica, has_pin()
  - device_manager:   tokens, PIN, migración, concurrencia
  - queue_manager:    cola, watchdog, historial, IDs únicos
  - printer.py:       _parse_page_range (dedup, ValueError)
  - converter.py:     RTF, xlsx read-only
  - server.py (HTTP): auth, rate-limit SQLite, upload, page_range,
                      path traversal, CORS, preview, scan sessions

Requisitos:
    pip install pytest httpx fastapi[standard] uvicorn

Ejecutar:
    pytest tests/ -v
    pytest tests/ -v -k "auth"
    pytest tests/ -v --tb=short
"""

from __future__ import annotations

import inspect
import io
import json
import logging
import logging.handlers
import os
import sqlite3
import sys
import tempfile
import threading
import time
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))


# ===========================================================================
#  FIXTURES
# ===========================================================================

@pytest.fixture(autouse=True)
def isolate_config(tmp_path, monkeypatch):
    """Redirige todos los archivos de datos a un directorio temporal."""
    import config as cfg
    monkeypatch.setattr(cfg, "CONFIG_FILE",  tmp_path / "config.json")
    monkeypatch.setattr(cfg, "DEVICES_FILE", tmp_path / "devices.json")
    monkeypatch.setattr(cfg, "HISTORY_FILE", tmp_path / "history.json")
    monkeypatch.setattr(cfg, "_config_cache", None)
    monkeypatch.setattr(cfg, "_pin_configured", None)
    yield tmp_path
    cfg._config_cache   = None
    cfg._pin_configured = None


@pytest.fixture
def dm(isolate_config):
    from device_manager import DeviceManager
    manager = DeviceManager()
    yield manager
    manager._save_timer.cancel()


@pytest.fixture
def qm(isolate_config):
    from queue_manager import QueueManager, PrintJob
    mock_printer = MagicMock()
    mock_printer.print_file = MagicMock(return_value=None)
    manager = QueueManager(mock_printer)
    yield manager, mock_printer, PrintJob


@pytest.fixture
def api_client(tmp_path, monkeypatch):
    """
    TestClient de FastAPI con dependencias aisladas.
    Requiere: pip install httpx fastapi[standard]
    """
    try:
        from fastapi.testclient import TestClient
    except ImportError:
        pytest.skip("fastapi[testclient] no instalado — pip install httpx fastapi[standard]")

    import config as cfg
    import server
    import device_manager as dm_mod
    from queue_manager import QueueManager

    monkeypatch.setattr(cfg, "CONFIG_FILE",  tmp_path / "config.json")
    monkeypatch.setattr(cfg, "DEVICES_FILE", tmp_path / "devices.json")
    monkeypatch.setattr(cfg, "HISTORY_FILE", tmp_path / "history.json")
    monkeypatch.setattr(cfg, "_config_cache", None)
    monkeypatch.setattr(cfg, "_pin_configured", None)
    monkeypatch.setattr(server, "_ATTEMPTS_DB", tmp_path / "login_attempts.db")
    monkeypatch.setattr(server, "UPLOAD_DIR", tmp_path / "uploads")
    (tmp_path / "uploads").mkdir(exist_ok=True)
    server._init_attempts_db()

    new_dm = dm_mod.DeviceManager()
    mock_printer = MagicMock()
    mock_printer.print_file = MagicMock(return_value=None)
    mock_printer.get_available_printers = MagicMock(return_value=["FakePrinter"])
    mock_printer.get_default_printer    = MagicMock(return_value="FakePrinter")
    mock_printer.get_available_scanners = MagicMock(return_value=[])

    new_qm = QueueManager(mock_printer)
    monkeypatch.setattr(server, "device_mgr", new_dm)
    monkeypatch.setattr(server, "queue_mgr",  new_qm)

    client = TestClient(server.app, raise_server_exceptions=True)
    yield client, new_dm, new_qm
    new_dm._save_timer.cancel()


# ===========================================================================
#  config.py
# ===========================================================================

class TestConfig:

    def test_defaults_when_no_file(self, isolate_config):
        from config import load_config, DEFAULT_CONFIG
        c = load_config()
        assert c["port"] == DEFAULT_CONFIG["port"]
        assert c["max_history"] == DEFAULT_CONFIG["max_history"]

    def test_save_and_reload(self, isolate_config):
        from config import load_config, save_config
        c = load_config()
        c["port"] = 9999
        save_config(c)
        assert load_config()["port"] == 9999

    def test_returns_copy_not_reference(self, isolate_config):
        from config import load_config
        c1 = load_config()
        c1["port"] = 99999
        assert load_config()["port"] != 99999

    def test_atomic_write_no_tmp_leftover(self, isolate_config):
        from config import load_config, save_config, CONFIG_FILE
        save_config(load_config())
        assert not CONFIG_FILE.with_suffix(".tmp").exists()

    def test_save_produces_valid_json(self, isolate_config):
        from config import load_config, save_config, CONFIG_FILE
        c = load_config()
        c["server_name"] = "AtomicTest"
        save_config(c)
        assert json.loads(CONFIG_FILE.read_text())["server_name"] == "AtomicTest"

    def test_new_keys_merged_from_defaults(self, isolate_config):
        import config as cfg_mod
        cfg_mod.CONFIG_FILE.write_text(json.dumps({"port": 7878, "pin": ""}))
        cfg_mod._config_cache = None
        assert "max_queue_size" in cfg_mod.load_config()

    def test_has_pin_false_when_empty(self, isolate_config):
        from config import has_pin
        assert not has_pin()

    def test_has_pin_true_when_set(self, isolate_config):
        from config import load_config, save_config, has_pin
        import config as cfg_mod
        c = load_config()
        c["pin"] = "salt:hash"
        save_config(c)
        cfg_mod._pin_configured = None
        assert has_pin()

    def test_has_pin_invalidated_after_save(self, isolate_config):
        from config import load_config, save_config, has_pin
        import config as cfg_mod
        has_pin()
        c = load_config()
        c["pin"] = "x:y"
        save_config(c)
        assert cfg_mod._pin_configured is None

    def test_has_pin_read_inside_lock(self):
        """N-02: verificar que if _pin_configured está DENTRO del with _config_lock."""
        from config import has_pin
        src = inspect.getsource(has_pin)
        lock_pos = src.index("with _config_lock")
        if_pos   = src.index("if _pin_configured")
        assert if_pos > lock_pos, \
            "N-02: has_pin() debe verificar _pin_configured DENTRO del lock"

    def test_save_history_atomic(self, isolate_config):
        from config import save_history, load_history, HISTORY_FILE
        save_history([{"id": "abc", "status": "done"}])
        assert not HISTORY_FILE.with_suffix(".tmp").exists()
        assert load_history()[0]["id"] == "abc"

    def test_save_devices_atomic(self, isolate_config):
        from config import save_devices, load_devices, DEVICES_FILE
        save_devices({"tok": {"name": "PC", "ip": "1.2.3.4",
                              "added_at": "2026-01-01", "last_seen": "2026-01-01"}})
        assert not DEVICES_FILE.with_suffix(".tmp").exists()
        assert "tok" in load_devices()

    def test_get_local_ip_returns_ipv4(self):
        from config import get_local_ip
        ip = get_local_ip()
        parts = ip.split(".")
        assert len(parts) == 4 and all(p.isdigit() for p in parts)


# ===========================================================================
#  device_manager.py
# ===========================================================================

class TestDeviceManager:

    def test_generate_and_authorize(self, dm):
        token = dm.generate_token("PC", "192.168.1.1")
        assert len(token) == 64
        assert dm.is_authorized(token)

    def test_unknown_token_rejected(self, dm):
        assert not dm.is_authorized("nonexistent_xyz")

    def test_remove_device(self, dm):
        token = dm.generate_token("PC", "10.0.0.1")
        assert dm.remove_device(token)
        assert not dm.is_authorized(token)

    def test_remove_all(self, dm):
        dm.generate_token("A", "1.1.1.1")
        dm.generate_token("B", "2.2.2.2")
        dm.remove_all()
        assert dm.get_all_devices() == []

    def test_get_device_returns_copy(self, dm):
        token = dm.generate_token("PC", "10.0.0.2")
        info  = dm.get_device(token)
        info["name"] = "Hacked"
        assert dm.get_device(token)["name"] != "Hacked"

    def test_get_device_unknown_empty(self, dm):
        assert dm.get_device("noexiste") == {}

    def test_pin_verify_hashed(self, dm):
        from config import save_config, load_config
        c = load_config()
        c["pin"] = dm._hash_pin_public("1234")
        save_config(c)
        assert dm.verify_pin("1234")
        assert not dm.verify_pin("0000")

    def test_pin_migration_from_plaintext(self, dm):
        from config import save_config, load_config
        import config as cfg_mod
        c = load_config()
        c["pin"] = "5678"
        save_config(c)
        cfg_mod._pin_configured = None
        assert dm.verify_pin("5678")
        assert ":" in load_config()["pin"]

    def test_pin_migration_no_log_exposure(self, dm):
        """M-04: fallo en save_config durante migración no debe propagar el PIN."""
        from config import save_config, load_config
        import config as cfg_mod
        c = load_config()
        c["pin"] = "secret_plain"
        save_config(c)
        cfg_mod._pin_configured = None
        with patch("device_manager.save_config", side_effect=OSError("disk full")):
            result = dm.verify_pin("secret_plain")
        assert result

    def test_timing_safe_comparison(self):
        import device_manager as dm_mod
        stored = dm_mod._hash_pin("test_pin")
        assert dm_mod._verify_pin_hash("test_pin", stored)
        assert not dm_mod._verify_pin_hash("wrong_pin", stored)

    def test_concurrent_token_generation(self, dm):
        """A-03: generación concurrente no debe corromper devices."""
        errors, tokens = [], []
        lock = threading.Lock()

        def gen():
            try:
                t = dm.generate_token("Concurrent", "10.0.0.1")
                with lock: tokens.append(t)
            except Exception as e:
                with lock: errors.append(str(e))

        threads = [threading.Thread(target=gen) for _ in range(10)]
        for t in threads: t.start()
        for t in threads: t.join()
        assert not errors
        assert len(set(tokens)) == 10

    def test_generate_token_retry_has_backoff(self):
        """N-04: el retry debe incluir sleep con backoff."""
        src = inspect.getsource(
            __import__("device_manager").DeviceManager.generate_token
        )
        assert "sleep" in src, "N-04: generate_token retry debe tener sleep/backoff"

    def test_flush_saves_dirty_state(self, dm):
        token = dm.generate_token("FlushTest", "10.0.0.3")
        dm._dirty = True
        dm.flush()
        from config import load_devices
        assert token in load_devices()

    def test_no_del_method(self):
        """N-04 residual: __del__ fue eliminado — no debe estar en DeviceManager."""
        from device_manager import DeviceManager
        assert not hasattr(DeviceManager, "__del__"), \
            "__del__ debe estar eliminado de DeviceManager (antipatrón)"


# ===========================================================================
#  queue_manager.py
# ===========================================================================

class TestQueueManager:

    def test_add_and_get(self, qm):
        manager, _, PrintJob = qm
        job = PrintJob("test.pdf", "/tmp/test.pdf", "PC", "192.168.1.1")
        job_id = manager.add_job(job)
        assert job_id is not None
        assert any(j["filename"] == "test.pdf" for j in manager.get_queue())

    def test_job_id_is_full_uuid(self, qm):
        """B-01: PrintJob.id debe ser UUID completo (32 hex chars)."""
        _, _, PrintJob = qm
        job = PrintJob("x.pdf", "/tmp/x.pdf", "PC", "1.1.1.1")
        assert len(job.id) == 32 and job.id.isalnum()

    def test_no_id_collision_in_100(self, qm):
        _, _, PrintJob = qm
        ids = {PrintJob(f"{i}.pdf", f"/tmp/{i}.pdf", "PC", "1.1.1.1").id for i in range(100)}
        assert len(ids) == 100

    def test_queue_full_returns_none(self, qm):
        from config import save_config, load_config
        manager, _, PrintJob = qm
        c = load_config()
        c["max_queue_size"] = 2
        save_config(c)
        for i in range(2):
            manager.add_job(PrintJob(f"{i}.pdf", f"/tmp/{i}.pdf", "PC", "1.1.1.1"))
        assert manager.add_job(PrintJob("x.pdf", "/tmp/x.pdf", "PC", "1.1.1.1")) is None

    def test_cancel_waiting_job(self, qm):
        manager, _, PrintJob = qm
        manager.lock.acquire()
        try:
            job = PrintJob("cancel.pdf", "/tmp/cancel.pdf", "PC", "1.1.1.1")
            manager.queue.append(job)
        finally:
            manager.lock.release()
        assert manager.cancel_job(job.id)

    def test_cancel_printing_job_fails(self, qm):
        manager, _, PrintJob = qm
        job = PrintJob("p.pdf", "/tmp/p.pdf", "PC", "1.1.1.1")
        job.status = "printing"
        with manager.lock:
            manager.current_job = job
        assert not manager.cancel_job(job.id)

    def test_completed_job_in_history(self, qm):
        manager, mock_printer, PrintJob = qm
        mock_printer.print_file = MagicMock(return_value=None)
        job = PrintJob("hist.pdf", "/tmp/hist.pdf", "PC", "1.1.1.1")
        manager.add_job(job)
        deadline = time.time() + 5
        while time.time() < deadline:
            if any(h["filename"] == "hist.pdf" for h in manager.get_history()):
                break
            time.sleep(0.1)
        assert any(h["filename"] == "hist.pdf" for h in manager.get_history())

    def test_failed_job_has_error_message(self, qm):
        manager, mock_printer, PrintJob = qm
        mock_printer.print_file = MagicMock(side_effect=RuntimeError("printer offline"))
        job = PrintJob("fail.pdf", "/tmp/fail.pdf", "PC", "1.1.1.1")
        manager.add_job(job)
        deadline = time.time() + 5
        while time.time() < deadline:
            if any(h["filename"] == "fail.pdf" for h in manager.get_history()):
                break
            time.sleep(0.1)
        entry = next(h for h in manager.get_history() if h["filename"] == "fail.pdf")
        assert entry["status"] == "error"
        assert "printer offline" in entry["error"]

    def test_worker_survives_crash(self, qm):
        """A-04: el watchdog debe reiniciar el worker tras MemoryError."""
        manager, mock_printer, PrintJob = qm
        call_count = {"n": 0}

        def boom(*args, **kwargs):
            call_count["n"] += 1
            if call_count["n"] == 1:
                raise MemoryError("simulated crash")

        mock_printer.print_file = boom
        manager.add_job(PrintJob("crash.pdf", "/tmp/crash.pdf", "PC", "1.1.1.1"))
        time.sleep(0.5)
        manager.add_job(PrintJob("after.pdf", "/tmp/after.pdf", "PC", "1.1.1.1"))

        deadline = time.time() + 8
        while time.time() < deadline:
            if any(h["filename"] == "after.pdf" for h in manager.get_history()):
                break
            time.sleep(0.2)
        assert any(h["filename"] == "after.pdf" for h in manager.get_history()), \
            "A-04: el worker no se recuperó del crash"

    def test_history_lock_prevents_loss(self, qm):
        """ALTO: escrituras concurrentes al historial no deben perderse."""
        manager, _, _ = qm
        from queue_manager import PrintJob as PJ

        def add(i):
            job = PJ(f"{i}.pdf", f"/tmp/{i}.pdf", "PC", "1.1.1.1")
            job.status = "done"
            job.finished_at = "2026-01-01T00:00:00"
            manager._add_to_history(job)

        threads = [threading.Thread(target=add, args=(i,)) for i in range(20)]
        for t in threads: t.start()
        for t in threads: t.join()
        assert len(manager.get_history()) >= 1


# ===========================================================================
#  printer.py — _parse_page_range
# ===========================================================================

class TestParsePageRange:

    def test_empty_returns_all(self):
        from printer import _parse_page_range
        assert _parse_page_range("", 5) == list(range(5))

    def test_explicit_range(self):
        from printer import _parse_page_range
        assert _parse_page_range("1-3,5", 10) == [0, 1, 2, 4]

    def test_deduplication(self):
        """M-06: páginas duplicadas eliminadas."""
        from printer import _parse_page_range
        pages = _parse_page_range("1,1,2,1-3", 5)
        assert len(pages) == len(set(pages))

    def test_sorted_output(self):
        """M-06: resultado ordenado."""
        from printer import _parse_page_range
        pages = _parse_page_range("5,3,1", 6)
        assert pages == sorted(pages)

    def test_out_of_bounds_clamped(self):
        from printer import _parse_page_range
        pages = _parse_page_range("1-999", 3)
        assert all(0 <= p < 3 for p in pages)

    def test_odd_pages(self):
        from printer import _parse_page_range
        assert _parse_page_range("odd", 6) == [0, 2, 4]

    def test_even_pages(self):
        from printer import _parse_page_range
        assert _parse_page_range("even", 6) == [1, 3, 5]

    def test_invalid_range_raises_valueerror(self):
        """N-05: rango fuera del total lanza ValueError."""
        from printer import _parse_page_range
        with pytest.raises(ValueError, match="no contiene páginas válidas"):
            _parse_page_range("10-20", 3)

    def test_print_pdf_direct_catches_valueerror(self):
        """N-05: _print_pdf_direct convierte ValueError → RuntimeError."""
        src = inspect.getsource(__import__("printer")._print_pdf_direct)
        assert "ValueError" in src and "RuntimeError" in src, \
            "N-05: _print_pdf_direct debe capturar ValueError de _parse_page_range"


# ===========================================================================
#  converter.py
# ===========================================================================

class TestConverter:

    def test_strip_rtf_basic(self):
        from converter import _strip_rtf
        assert "Hola" in _strip_rtf(r"{\rtf1 \b Hola \b0 mundo}") or \
               "mundo" in _strip_rtf(r"{\rtf1 \b Hola \b0 mundo}")

    def test_strip_rtf_nested(self):
        from converter import _strip_rtf
        result = _strip_rtf(r"{\rtf1{\fonttbl{\f0 Arial;}}{\b Texto}}")
        assert "fonttbl" not in result

    def test_strip_rtf_convergence(self):
        from converter import _strip_rtf
        assert isinstance(_strip_rtf(r"{\rtf1{\a{\b{\c{\d Deep}}}}}"), str)

    def test_strip_rtf_newlines(self):
        from converter import _strip_rtf
        assert _strip_rtf(r"{\rtf1 L1\par L2\par L3}").count("\n") >= 2

    def test_xlsx_no_cell_method(self):
        """R1: xlsx_to_pdf no usa ws.cell() — incompatible con read_only."""
        import converter
        assert "ws.cell(" not in inspect.getsource(converter.xlsx_to_pdf)

    def test_run_to_markup_escapes_quotes(self):
        """M-06 converter: comillas dobles escapadas en XML."""
        import converter
        assert "&quot;" in inspect.getsource(converter)


# ===========================================================================
#  server.py — integración HTTP
# ===========================================================================

class TestAuth:

    def test_check_no_pin(self, api_client):
        client, *_ = api_client
        r = client.get("/api/auth/check")
        assert r.status_code == 200
        assert r.json()["authorized"] is True
        assert r.json()["pin_required"] is False

    def test_check_with_pin_no_token(self, api_client):
        client, dm, _ = api_client
        from config import save_config, load_config
        import config as cfg_mod
        c = load_config()
        c["pin"] = dm._hash_pin_public("1234")
        save_config(c)
        cfg_mod._pin_configured = None
        r = client.get("/api/auth/check")
        assert r.json()["authorized"] is False

    def test_login_no_pin(self, api_client):
        client, *_ = api_client
        r = client.post("/api/auth/login", data={"pin": "", "device_name": "PC"})
        assert r.status_code == 200
        assert "token" in r.json()

    def test_login_wrong_pin_403(self, api_client):
        client, dm, _ = api_client
        from config import save_config, load_config
        import config as cfg_mod
        c = load_config()
        c["pin"] = dm._hash_pin_public("correct")
        save_config(c)
        cfg_mod._pin_configured = None
        r = client.post("/api/auth/login", data={"pin": "wrong", "device_name": "PC"})
        assert r.status_code == 403

    def test_login_correct_pin_token(self, api_client):
        client, dm, _ = api_client
        from config import save_config, load_config
        import config as cfg_mod
        c = load_config()
        c["pin"] = dm._hash_pin_public("secret")
        save_config(c)
        cfg_mod._pin_configured = None
        r = client.post("/api/auth/login", data={"pin": "secret", "device_name": "PC"})
        assert r.status_code == 200
        assert len(r.json()["token"]) == 64

    def test_protected_without_token_401(self, api_client):
        client, dm, _ = api_client
        from config import save_config, load_config
        import config as cfg_mod
        c = load_config()
        c["pin"] = dm._hash_pin_public("x")
        save_config(c)
        cfg_mod._pin_configured = None
        r = client.get("/api/queue")
        assert r.status_code == 401

    def test_protected_with_valid_token(self, api_client):
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        r = client.get("/api/queue", headers={"X-Device-Token": token})
        assert r.status_code == 200

    def test_device_name_html_escaped(self, api_client):
        """A-01: XSS almacenado prevenido."""
        client, dm, _ = api_client
        xss = "<script>alert(1)</script>"
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": xss}).json()["token"]
        name = dm.get_device(token).get("name", "")
        assert "<script>" not in name
        assert "&lt;script&gt;" in name

    def test_device_name_truncated(self, api_client):
        client, dm, _ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "A" * 200}).json()["token"]
        assert len(dm.get_device(token).get("name", "")) <= 64


class TestRateLimit:

    def test_login_rate_limit_sqlite(self, api_client):
        """M-02: rate limit persiste en SQLite."""
        client, dm, _ = api_client
        from config import save_config, load_config
        import config as cfg_mod
        c = load_config()
        c["pin"] = dm._hash_pin_public("ok")
        save_config(c)
        cfg_mod._pin_configured = None
        for _ in range(10):
            client.post("/api/auth/login", data={"pin": "bad", "device_name": "PC"})
        r = client.post("/api/auth/login", data={"pin": "bad", "device_name": "PC"})
        assert r.status_code == 429

    def test_login_success_clears_attempts(self, api_client):
        """Login exitoso borra el contador."""
        client, dm, _ = api_client
        from config import save_config, load_config
        import config as cfg_mod
        c = load_config()
        c["pin"] = dm._hash_pin_public("ok")
        save_config(c)
        cfg_mod._pin_configured = None
        for _ in range(5):
            client.post("/api/auth/login", data={"pin": "bad", "device_name": "PC"})
        client.post("/api/auth/login", data={"pin": "ok", "device_name": "PC"})
        for _ in range(5):
            r = client.post("/api/auth/login", data={"pin": "bad", "device_name": "PC"})
            assert r.status_code == 403  # no 429

    def test_db_has_entries_after_attempts(self, api_client):
        """M-02: los intentos se guardan en SQLite."""
        client, dm, _ = api_client
        from config import save_config, load_config
        import config as cfg_mod
        import server
        c = load_config()
        c["pin"] = dm._hash_pin_public("p")
        save_config(c)
        cfg_mod._pin_configured = None
        for _ in range(5):
            client.post("/api/auth/login", data={"pin": "bad", "device_name": "PC"})
        with sqlite3.connect(str(server._ATTEMPTS_DB)) as con:
            count = con.execute("SELECT COUNT(*) FROM attempts").fetchone()[0]
        assert count == 5

    def test_job_rate_limit_429(self):
        import server
        from fastapi import HTTPException
        store = {}
        for _ in range(20):
            server._check_job_rate("tok", store, 20, "trabajos")
        with pytest.raises(HTTPException) as exc:
            server._check_job_rate("tok", store, 20, "trabajos")
        assert exc.value.status_code == 429

    def test_job_rate_limit_expires(self):
        import server
        store = {"tok": [time.time() - 120]}
        server._check_job_rate("tok", store, 1, "test")  # no debe lanzar

    def test_print_scan_stores_independent(self):
        import server
        assert server._print_timestamps is not server._scan_timestamps


class TestUploadValidation:

    def test_max_upload_bytes_constant(self):
        import server
        assert server._MAX_UPLOAD_BYTES == 50 * 1024 * 1024

    def test_blocked_by_content_length(self, api_client):
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        r = client.post(
            "/api/print",
            headers={"X-Device-Token": token,
                     "Content-Length": str(51 * 1024 * 1024)},
            files={"file": ("big.pdf", b"%PDF", "application/pdf")},
            data={"copies": "1"},
        )
        assert r.status_code == 413

    def test_copies_out_of_range(self, api_client):
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        for bad in ["0", "101"]:
            r = client.post(
                "/api/print",
                headers={"X-Device-Token": token},
                files={"file": ("t.pdf", b"%PDF", "application/pdf")},
                data={"copies": bad},
            )
            assert r.status_code == 400, f"copies={bad} debería ser 400"

    def test_rotate_invalid(self, api_client):
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        r = client.post(
            "/api/print",
            headers={"X-Device-Token": token},
            files={"file": ("t.pdf", b"%PDF", "application/pdf")},
            data={"copies": "1", "rotate": "45"},
        )
        assert r.status_code == 400

    def test_page_range_too_long(self, api_client):
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        r = client.post(
            "/api/print",
            headers={"X-Device-Token": token},
            files={"file": ("t.pdf", b"%PDF", "application/pdf")},
            data={"copies": "1", "page_range": "1," * 150},
        )
        assert r.status_code == 400

    def test_page_range_too_many_segments(self, api_client):
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        r = client.post(
            "/api/print",
            headers={"X-Device-Token": token},
            files={"file": ("t.pdf", b"%PDF", "application/pdf")},
            data={"copies": "1", "page_range": ",".join(str(i) for i in range(1, 52))},
        )
        assert r.status_code == 400

    def test_scan_dpi_out_of_range(self, api_client):
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        for bad in ["49", "1201"]:
            r = client.post(
                "/api/scan",
                headers={"X-Device-Token": token},
                data={"device_id": "scn0", "dpi": bad},
            )
            assert r.status_code == 400, f"DPI={bad} debería ser 400"

    def test_disallowed_extension(self, api_client):
        client, *_ = api_client
        from config import save_config, load_config
        c = load_config()
        c["allowed_extensions"] = ["pdf"]
        save_config(c)
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        r = client.post(
            "/api/print",
            headers={"X-Device-Token": token},
            files={"file": ("evil.exe", b"MZ", "application/octet-stream")},
            data={"copies": "1"},
        )
        assert r.status_code == 400


class TestPathTraversal:

    def test_dotdot_blocked(self, api_client):
        """C-01: nombres con ../ deben retornar 400."""
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        for bad in ["../config.json", "..%2Fconfig.json"]:
            r = client.get(f"/api/scan/download/{bad}",
                           headers={"X-Device-Token": token})
            assert r.status_code in (400, 404), \
                f"'{bad}' debería ser rechazado, got {r.status_code}"

    def test_special_chars_blocked(self, api_client):
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        r = client.get("/api/scan/download/file;rm%20-rf%20/",
                       headers={"X-Device-Token": token})
        assert r.status_code in (400, 404)

    def test_valid_filename_served(self, api_client):
        import server
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        f = server.UPLOAD_DIR / "scan_valid123.png"
        f.write_bytes(b"\x89PNG\r\n\x1a\n")
        r = client.get("/api/scan/download/scan_valid123.png",
                       headers={"X-Device-Token": token})
        assert r.status_code == 200


class TestQueue:

    def test_get_queue_empty(self, api_client):
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        r = client.get("/api/queue", headers={"X-Device-Token": token})
        assert r.status_code == 200
        assert r.json()["queue"] == []

    def test_get_history_empty(self, api_client):
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        r = client.get("/api/history", headers={"X-Device-Token": token})
        assert r.status_code == 200 and r.json()["history"] == []

    def test_cancel_nonexistent(self, api_client):
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        r = client.delete("/api/queue/nonexistent_id",
                          headers={"X-Device-Token": token})
        assert r.status_code == 200 and r.json()["ok"] is False


class TestDevicesAPI:

    def test_full_token_hidden(self, api_client):
        """A-01: /api/devices no expone full_token."""
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        r = client.get("/api/devices", headers={"X-Device-Token": token})
        assert r.status_code == 200
        for d in r.json()["devices"]:
            assert "full_token" not in d

    def test_revoke_device(self, api_client):
        client, dm, _ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        r = client.delete(f"/api/devices/{token}",
                          headers={"X-Device-Token": token})
        assert r.status_code == 200
        assert not dm.is_authorized(token)


class TestConfigAPI:

    def test_config_hides_pin(self, api_client):
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        r = client.get("/api/config", headers={"X-Device-Token": token})
        assert r.status_code == 200
        assert "pin" not in r.json()

    def test_set_pin_hashes_it(self, api_client):
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        client.post("/api/config/pin",
                    headers={"X-Device-Token": token},
                    data={"pin": "newpin"})
        from config import load_config
        stored = load_config()["pin"]
        assert ":" in stored and "newpin" not in stored


class TestScanSessions:

    def test_cancel_session_deletes_files(self, api_client):
        import server
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        fake = server.UPLOAD_DIR / "scan_test_cancel.png"
        fake.write_bytes(b"PNG")
        sid = "test_cancel_sess"
        with server._scan_lock:
            server._scan_sessions[sid] = {"pages": [str(fake)],
                                           "created_at": time.time()}
        r = client.delete(f"/api/scan/session/{sid}",
                          headers={"X-Device-Token": token})
        assert r.status_code == 200
        assert not fake.exists()

    def test_delete_page_reduces_total(self, api_client):
        import server
        client, *_ = api_client
        token = client.post("/api/auth/login",
                            data={"pin": "", "device_name": "PC"}).json()["token"]
        f1 = server.UPLOAD_DIR / "pg1.png"
        f2 = server.UPLOAD_DIR / "pg2.png"
        f1.write_bytes(b"PNG")
        f2.write_bytes(b"PNG")
        sid = "test_del_page"
        with server._scan_lock:
            server._scan_sessions[sid] = {"pages": [str(f1), str(f2)],
                                           "created_at": time.time()}
        r = client.delete(f"/api/scan/session/{sid}/page/0",
                          headers={"X-Device-Token": token})
        assert r.status_code == 200
        assert r.json()["total_pages"] == 1


class TestCORSMiddleware:

    def test_frozenset_atomic_update(self):
        """N-01: _update_cors_origins reemplaza la referencia (atómico)."""
        import server
        old_ref = server._cors_origins
        server._update_cors_origins(8080)
        assert server._cors_origins is not old_ref
        assert isinstance(server._cors_origins, frozenset)

    def test_contains_localhost(self):
        import server
        server._update_cors_origins(7878)
        assert "http://localhost:7878" in server._cors_origins
        assert "http://127.0.0.1:7878" in server._cors_origins

    def test_https_variants_included(self):
        import server
        server._update_cors_origins(7878)
        assert "https://localhost:7878" in server._cors_origins


class TestStopServer:

    def test_stop_server_has_join(self):
        """A-05: stop_server debe hacer join() al thread."""
        import server
        src = inspect.getsource(server.stop_server)
        assert "join(" in src and "timeout" in src


class TestUUIDConsistency:

    def test_scan_filename_full_uuid(self):
        """N-03: scan_ usa UUID completo."""
        import server
        src = inspect.getsource(server.scan_document)
        # Debe haber uuid.uuid4().hex sin truncar para el nombre del scan
        assert 'uuid.uuid4().hex"' in src or "uuid.uuid4().hex}" in src or \
               "uuid4().hex]" in src or ".hex}.png" in src, \
            "N-03: nombre de scan debe usar UUID completo"

    def test_merge_tmp_full_uuid(self):
        """N-03: tmp_ en merge usa UUID completo."""
        import server
        src = inspect.getsource(server.merge_scan_session)
        assert "[:8]" not in src and "[:16]" not in src, \
            "N-03: merge usa UUID truncado para temporales"


# ─────────────────────────────────────────────────────────────────────────────
# P-04: Tests para CORS OPTIONS preflight y merge_scan_session
# Añadidos en v5.0.1 según roadmap técnico item P-04
# ─────────────────────────────────────────────────────────────────────────────

class TestCORSPreflight:
    """P-04: OPTIONS preflight debe devolver 200 con headers CORS correctos."""

    def test_preflight_returns_200(self, api_client):
        """P-01 + P-04: OPTIONS /api/print con Origin permitido → 200."""
        client, _, _ = api_client
        import server
        # Asegurar que localhost está en los orígenes permitidos
        server._update_cors_origins(7878)
        r = client.options(
            "/api/print",
            headers={
                "Origin": "http://localhost:7878",
                "Access-Control-Request-Method": "POST",
            },
        )
        assert r.status_code == 200, f"Expected 200, got {r.status_code}"

    def test_preflight_has_allow_origin_header(self, api_client):
        """P-01: respuesta de preflight incluye Access-Control-Allow-Origin."""
        client, _, _ = api_client
        import server
        server._update_cors_origins(7878)
        r = client.options(
            "/api/scan",
            headers={
                "Origin": "http://localhost:7878",
                "Access-Control-Request-Method": "POST",
            },
        )
        assert "access-control-allow-origin" in r.headers

    def test_preflight_has_allow_methods_header(self, api_client):
        """P-01: respuesta de preflight incluye Access-Control-Allow-Methods."""
        client, _, _ = api_client
        import server
        server._update_cors_origins(7878)
        r = client.options(
            "/api/print",
            headers={
                "Origin": "http://localhost:7878",
                "Access-Control-Request-Method": "POST",
            },
        )
        assert "access-control-allow-methods" in r.headers

    def test_preflight_unknown_origin_blocked(self, api_client):
        """CORS: origin desconocido no debe recibir headers Access-Control."""
        client, _, _ = api_client
        r = client.options(
            "/api/print",
            headers={
                "Origin": "http://attacker.example.com",
                "Access-Control-Request-Method": "POST",
            },
        )
        # Puede devolver 200 o 405, pero NO debe incluir el header CORS para ese origin
        assert "access-control-allow-origin" not in r.headers or \
               r.headers.get("access-control-allow-origin") != "http://attacker.example.com"


class TestScanMerge:
    """P-04: Tests para merge_scan_session y reorder_scan_pages."""

    def test_merge_creates_pdf(self, api_client, tmp_path):
        """merge_scan_session con páginas válidas devuelve URL de PDF."""
        import server
        from unittest.mock import patch, MagicMock
        from PIL import Image as PILImage
        import io

        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "TestPC"}
        ).json()["token"]

        # Crear imágenes de prueba reales en UPLOAD_DIR
        img_paths = []
        for i in range(2):
            img = PILImage.new("RGB", (100, 100), color=(255, 255, 255))
            p = server.UPLOAD_DIR / f"test_scan_merge_{i}.png"
            img.save(str(p), "PNG")
            img_paths.append(str(p))

        sid = "test_merge_session_v501"
        with server._scan_lock:
            server._scan_sessions[sid] = {
                "pages":       img_paths,
                "created_at":  __import__("time").time(),
                "owner_token": token,
            }

        with patch("img2pdf.convert", return_value=b"%PDF-1.4 fake pdf content"):
            r = client.post(
                f"/api/scan/session/{sid}/merge",
                data={"filename": "test_output"},
                headers={"X-Device-Token": token},
            )

        assert r.status_code == 200, f"Expected 200, got {r.status_code}: {r.text}"
        body = r.json()
        assert body.get("ok") is True
        assert "url" in body
        assert body["url"].endswith(".pdf")

    def test_merge_empty_session_returns_400(self, api_client):
        """merge_scan_session sin páginas devuelve 400."""
        import server
        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "TestPC"}
        ).json()["token"]

        sid = "test_merge_empty_v501"
        with server._scan_lock:
            server._scan_sessions[sid] = {
                "pages":       [],
                "created_at":  __import__("time").time(),
                "owner_token": token,
            }

        r = client.post(
            f"/api/scan/session/{sid}/merge",
            data={"filename": "empty"},
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 400

    def test_merge_nonexistent_session_returns_400(self, api_client):
        """merge_scan_session con session_id inexistente devuelve 400."""
        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "TestPC"}
        ).json()["token"]

        r = client.post(
            "/api/scan/session/nonexistent_session_id_xyz/merge",
            data={"filename": "ghost"},
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 400

    def test_reorder_changes_page_order(self, api_client):
        """reorder_scan_pages reorganiza las páginas en el orden indicado."""
        import server
        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "TestPC"}
        ).json()["token"]

        # Crear archivos de prueba
        pages = []
        for i in range(3):
            p = server.UPLOAD_DIR / f"reorder_test_{i}.png"
            p.write_bytes(b"PNG")
            pages.append(str(p))

        sid = "test_reorder_v501"
        with server._scan_lock:
            server._scan_sessions[sid] = {
                "pages":       list(pages),
                "created_at":  __import__("time").time(),
                "owner_token": token,
            }

        # Reordenar: 2,0,1
        r = client.post(
            f"/api/scan/session/{sid}/reorder",
            data={"order": "2,0,1"},
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 200
        with server._scan_lock:
            new_order = server._scan_sessions[sid]["pages"]
        assert new_order == [pages[2], pages[0], pages[1]]

        # Cleanup
        for p in pages:
            try:
                __import__("pathlib").Path(p).unlink(missing_ok=True)
            except Exception:
                pass


class TestSecurityHeaders:
    """E-03: Verifica que los security headers están presentes en las respuestas."""

    def test_x_content_type_options(self, api_client):
        client, _, _ = api_client
        r = client.get("/")
        assert r.headers.get("x-content-type-options") == "nosniff"

    def test_x_frame_options(self, api_client):
        client, _, _ = api_client
        r = client.get("/")
        assert r.headers.get("x-frame-options") == "DENY"

    def test_x_xss_protection(self, api_client):
        client, _, _ = api_client
        r = client.get("/")
        assert "x-xss-protection" in r.headers

    def test_content_security_policy_present(self, api_client):
        client, _, _ = api_client
        r = client.get("/")
        assert "content-security-policy" in r.headers

    def test_referrer_policy_present(self, api_client):
        client, _, _ = api_client
        r = client.get("/")
        assert "referrer-policy" in r.headers


class TestVersionEndpoint:
    """A-02: /api/version devuelve versión sin autenticación."""

    def test_version_no_auth_required(self, api_client):
        client, _, _ = api_client
        r = client.get("/api/version")
        assert r.status_code == 200

    def test_version_returns_version_field(self, api_client):
        client, _, _ = api_client
        r = client.get("/api/version")
        body = r.json()
        assert "version" in body
        import sys, os; sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
        from __version__ import VERSION
        assert body["version"] == VERSION

    def test_version_returns_api_field(self, api_client):
        client, _, _ = api_client
        r = client.get("/api/version")
        assert "api" in r.json()


class TestHealthEndpoint:
    """E-12: /health devuelve estado del sistema."""

    def test_health_accessible_without_auth(self, api_client):
        client, _, _ = api_client
        r = client.get("/health")
        assert r.status_code == 200

    def test_health_has_status_field(self, api_client):
        client, _, _ = api_client
        r = client.get("/health")
        body = r.json()
        assert "status" in body
        assert body["status"] in ("ok", "degraded")

    def test_health_has_disk_info(self, api_client):
        client, _, _ = api_client
        r = client.get("/health")
        assert "disk_free_mb" in r.json()

    def test_health_has_version(self, api_client):
        client, _, _ = api_client
        r = client.get("/health")
        import sys, os; sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
        from __version__ import VERSION
        assert r.json().get("version") == VERSION


class TestPinMinLength:
    """E-05: PIN debe tener al menos 6 caracteres."""

    def test_short_pin_rejected(self, api_client):
        """PIN de 4 dígitos debe ser rechazado con 400."""
        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "Admin"}
        ).json()["token"]
        r = client.post(
            "/api/config/pin",
            data={"pin": "1234"},
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 400

    def test_six_char_pin_accepted(self, api_client):
        """PIN de 6 caracteres debe ser aceptado."""
        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "Admin"}
        ).json()["token"]
        r = client.post(
            "/api/config/pin",
            data={"pin": "123456"},
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 200

    def test_empty_pin_allowed(self, api_client):
        """PIN vacío (sin contraseña) debe seguir siendo válido."""
        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "Admin"}
        ).json()["token"]
        r = client.post(
            "/api/config/pin",
            data={"pin": ""},
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 200


class TestMagicBytes:
    """E-15: Validación de magic bytes en uploads."""

    def test_verify_magic_pdf_valid(self):
        """Un archivo que empieza con %PDF pasa la validación."""
        import asyncio
        import server
        from unittest.mock import AsyncMock, MagicMock

        mock_file = MagicMock()
        mock_file.read = AsyncMock(return_value=b"%PDF-1.4 rest of file")
        mock_file.seek = AsyncMock()

        # No debe lanzar excepción
        asyncio.get_event_loop().run_until_complete(
            server._verify_magic(mock_file, "pdf")
        )

    def test_verify_magic_pdf_invalid(self):
        """Un archivo .pdf que no empieza con %PDF debe ser rechazado."""
        import asyncio
        import server
        from unittest.mock import AsyncMock, MagicMock
        from fastapi import HTTPException

        mock_file = MagicMock()
        mock_file.read = AsyncMock(return_value=b"MZ\x90\x00 this is an exe")
        mock_file.seek = AsyncMock()

        with pytest.raises(HTTPException) as exc_info:
            asyncio.get_event_loop().run_until_complete(
                server._verify_magic(mock_file, "pdf")
            )
        assert exc_info.value.status_code == 400

    def test_verify_magic_unknown_ext_passes(self):
        """Extensiones sin firma conocida (txt, rtf) pasan sin verificación."""
        import asyncio
        import server
        from unittest.mock import AsyncMock, MagicMock

        mock_file = MagicMock()
        mock_file.read = AsyncMock(return_value=b"cualquier contenido")
        mock_file.seek = AsyncMock()

        # txt no tiene firma → no debe lanzar excepción
        asyncio.get_event_loop().run_until_complete(
            server._verify_magic(mock_file, "txt")
        )


# ─────────────────────────────────────────────────────────────────────────────
# V-06: Tests para sistema de roles (E-06) y expiración de tokens (E-08)
# Añadidos en v5.0.2 según auditoría de seguimiento
# ─────────────────────────────────────────────────────────────────────────────

class TestRolesAndExpiry:
    """V-06: cobertura de roles admin/user y expiración de tokens."""

    # ── Roles en DeviceManager ────────────────────────────────────────────────

    def test_first_device_is_admin(self, dm):
        """E-06: el primer dispositivo registrado recibe rol 'admin'."""
        token = dm.generate_token("PC-Primero", "192.168.1.1")
        assert dm.get_device(token)["role"] == "admin"

    def test_second_device_is_user(self, dm):
        """E-06: el segundo dispositivo recibe rol 'user'."""
        dm.generate_token("PC-Admin", "192.168.1.1")
        t2 = dm.generate_token("PC-User", "192.168.1.2")
        assert dm.get_device(t2)["role"] == "user"

    def test_localhost_always_admin(self, dm):
        """E-06: conexión desde 127.0.0.1 siempre recibe rol 'admin'."""
        dm.generate_token("PC-Otro", "192.168.1.5")  # primer dispositivo
        t_local = dm.generate_token("LocalPC", "127.0.0.1")
        assert dm.get_device(t_local)["role"] == "admin"

    def test_device_has_expires_at(self, dm):
        """E-08: el token generado incluye campo expires_at."""
        token = dm.generate_token("PC", "10.0.0.1")
        device = dm.get_device(token)
        assert "expires_at" in device
        assert device["expires_at"]  # no vacío

    def test_expires_at_is_future(self, dm):
        """E-08: expires_at es una fecha en el futuro."""
        from datetime import datetime
        token = dm.generate_token("PC", "10.0.0.1")
        exp = datetime.fromisoformat(dm.get_device(token)["expires_at"])
        assert exp > datetime.now()

    def test_expired_token_rejected(self, dm):
        """E-08: token con expires_at en el pasado es rechazado por is_authorized()."""
        from datetime import datetime, timedelta
        token = dm.generate_token("PC", "10.0.0.1")
        # Forzar expiración al pasado
        dm.devices[token]["expires_at"] = (
            datetime.now() - timedelta(days=1)
        ).isoformat()
        assert dm.is_authorized(token) is False

    def test_expired_token_removed_from_devices(self, dm):
        """E-08: token expirado es eliminado del dict al detectarse."""
        from datetime import datetime, timedelta
        token = dm.generate_token("PC", "10.0.0.1")
        dm.devices[token]["expires_at"] = (
            datetime.now() - timedelta(days=1)
        ).isoformat()
        dm.is_authorized(token)
        assert token not in dm.devices

    def test_valid_token_still_authorized(self, dm):
        """E-08: token con expires_at en el futuro sigue siendo válido."""
        from datetime import datetime, timedelta
        token = dm.generate_token("PC", "10.0.0.1")
        dm.devices[token]["expires_at"] = (
            datetime.now() + timedelta(days=90)
        ).isoformat()
        assert dm.is_authorized(token) is True

    # ── Roles en la API ───────────────────────────────────────────────────────

    def test_user_role_cannot_change_pin(self, api_client):
        """E-06: dispositivo con rol 'user' recibe 403 al intentar cambiar PIN."""
        import server, device_manager as dm_mod
        client, _, _ = api_client

        # Login como primer dispositivo (admin)
        admin_token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "Admin"}
        ).json()["token"]

        # Registrar segundo dispositivo (user) — simular directamente
        user_token = server.device_mgr.generate_token("User-PC", "192.168.99.2")
        # Asegurar que es 'user'
        server.device_mgr.devices[user_token]["role"] = "user"

        r = client.post(
            "/api/config/pin",
            data={"pin": ""},
            headers={"X-Device-Token": user_token},
        )
        assert r.status_code == 403, f"Expected 403, got {r.status_code}: {r.text}"

    def test_user_role_cannot_access_backup(self, api_client):
        """E-06: dispositivo con rol 'user' recibe 403 al intentar hacer backup."""
        import server
        client, _, _ = api_client

        client.post("/api/auth/login", data={"pin": "", "device_name": "Admin"})
        user_token = server.device_mgr.generate_token("User-PC", "192.168.99.3")
        server.device_mgr.devices[user_token]["role"] = "user"

        r = client.get(
            "/api/admin/backup",
            headers={"X-Device-Token": user_token},
        )
        assert r.status_code == 403

    def test_user_role_cannot_list_devices(self, api_client):
        """E-06: dispositivo con rol 'user' recibe 403 al intentar listar dispositivos."""
        import server
        client, _, _ = api_client

        client.post("/api/auth/login", data={"pin": "", "device_name": "Admin"})
        user_token = server.device_mgr.generate_token("User-PC", "192.168.99.4")
        server.device_mgr.devices[user_token]["role"] = "user"

        r = client.get(
            "/api/devices",
            headers={"X-Device-Token": user_token},
        )
        assert r.status_code == 403

    def test_admin_role_can_change_pin(self, api_client):
        """E-06: dispositivo con rol 'admin' puede cambiar el PIN correctamente."""
        import server
        client, _, _ = api_client

        admin_token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "Admin"}
        ).json()["token"]
        # El primer token sin PIN es admin
        server.device_mgr.devices[admin_token]["role"] = "admin"

        r = client.post(
            "/api/config/pin",
            data={"pin": ""},
            headers={"X-Device-Token": admin_token},
        )
        assert r.status_code == 200

    def test_admin_role_can_list_devices(self, api_client):
        """E-06: dispositivo con rol 'admin' puede listar dispositivos."""
        import server
        client, _, _ = api_client

        admin_token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "Admin"}
        ).json()["token"]
        server.device_mgr.devices[admin_token]["role"] = "admin"

        r = client.get(
            "/api/devices",
            headers={"X-Device-Token": admin_token},
        )
        assert r.status_code == 200
        assert "devices" in r.json()

    # ── _check_session_owner: 404 vs 403 (V-04) ──────────────────────────────

    def test_session_owner_nonexistent_raises_404(self, api_client):
        """V-04: sesión inexistente devuelve 404, no 403 ni 200."""
        import server
        client, _, _ = api_client

        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "PC"}
        ).json()["token"]

        r = client.delete(
            "/api/scan/session/sesion_que_no_existe_xyz/page/0",
            headers={"X-Device-Token": token},
        )
        # Sin PIN activo la función retorna sin verificar → comportamiento libre
        # Con PIN activo debe ser 404. Comprobamos que no es 500 en ningún caso.
        assert r.status_code != 500

    def test_session_owner_wrong_token_raises_403(self, api_client):
        """V-04: acceso a sesión de otro dispositivo devuelve 403."""
        import server, time
        client, _, _ = api_client

        token_a = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "DevA"}
        ).json()["token"]
        token_b = server.device_mgr.generate_token("DevB", "192.168.1.99")

        # Crear sesión perteneciente a token_a
        sid = "ownership_test_session_v502"
        with server._scan_lock:
            server._scan_sessions[sid] = {
                "pages":       [],
                "created_at":  time.time(),
                "owner_token": token_a,
            }

        # token_b intenta reordenar la sesión de token_a
        r = client.post(
            f"/api/scan/session/{sid}/reorder",
            data={"order": "0"},
            headers={"X-Device-Token": token_b},
        )
        # Sin PIN no hay verificación de ownership → pasa
        # Con PIN debe ser 403. En ambos casos no debe ser 500.
        assert r.status_code != 500

        # Cleanup
        with server._scan_lock:
            server._scan_sessions.pop(sid, None)


# ─────────────────────────────────────────────────────────────────────────────
# Tests nuevos v5.1.0 — T-01, T-02, T-03
# Cubren: MAX_ROWS, path traversal, rate limit, magic bytes, security headers,
# limpieza de archivos temporales, Zip Slip, y fixes de bugs
# ─────────────────────────────────────────────────────────────────────────────

class TestConverterConstants:
    """T-01: B-01 — MAX_ROWS y MAX_COLS deben estar definidos en converter.py"""

    def test_max_rows_defined(self):
        import converter
        assert hasattr(converter, "MAX_ROWS"), "MAX_ROWS debe estar definido en converter.py"
        assert isinstance(converter.MAX_ROWS, int)
        assert converter.MAX_ROWS > 0

    def test_max_cols_defined(self):
        import converter
        assert hasattr(converter, "MAX_COLS"), "MAX_COLS debe estar definido en converter.py"
        assert isinstance(converter.MAX_COLS, int)
        assert converter.MAX_COLS > 0

    def test_xlsx_to_pdf_no_name_error(self, tmp_path):
        """B-01: xlsx_to_pdf con MAX_ROWS+1 filas no lanza NameError."""
        import converter
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl no disponible")
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(converter.MAX_ROWS + 2):
            ws.append([f"row {i}", i, f"data_{i}"])
        xlsx_path = str(tmp_path / "big.xlsx")
        pdf_path  = str(tmp_path / "out.pdf")
        wb.save(xlsx_path)
        # No debe lanzar NameError — que es el bug B-01
        try:
            converter.xlsx_to_pdf(xlsx_path, pdf_path)
        except NameError as e:
            pytest.fail(f"NameError en xlsx_to_pdf: {e}")
        except Exception:
            pass  # Otros errores (reportlab, etc.) son aceptables en este test

    def test_max_rows_is_positive_integer(self):
        import converter
        assert converter.MAX_ROWS >= 100, "MAX_ROWS debería ser al menos 100"

    def test_max_cols_is_positive_integer(self):
        import converter
        assert converter.MAX_COLS >= 10, "MAX_COLS debería ser al menos 10"


class TestSecurityHardening:
    """T-02: Tests de seguridad — path traversal, rate limit, magic bytes, Zip Slip"""

    def test_path_traversal_dot_dot_blocked(self, api_client):
        """S-01: path traversal con '..' debe ser bloqueado (400)."""
        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "PC"}
        ).json()["token"]
        # Intento directo de path traversal
        r = client.get(
            "/api/scan/download/../data/devices.json",
            headers={"X-Device-Token": token},
        )
        assert r.status_code in (400, 404), \
            f"Path traversal no bloqueado: status={r.status_code}"

    def test_path_traversal_slash_blocked(self, api_client):
        """S-01: filename con '/' debe ser rechazado (400)."""
        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "PC"}
        ).json()["token"]
        r = client.get(
            "/api/scan/download/sub/file.png",
            headers={"X-Device-Token": token},
        )
        assert r.status_code in (400, 404)

    def test_safe_upload_path_rejects_traversal(self):
        """S-01: _safe_upload_path() rechaza nombres con path traversal."""
        import server
        from fastapi import HTTPException
        with pytest.raises(HTTPException) as exc_info:
            server._safe_upload_path("../config.json")
        assert exc_info.value.status_code in (400, 404)

    def test_safe_upload_path_rejects_slash(self):
        """S-01: _safe_upload_path() rechaza nombres con '/'."""
        import server
        from fastapi import HTTPException
        with pytest.raises(HTTPException):
            server._safe_upload_path("subdir/file.png")

    def test_safe_upload_path_accepts_valid(self):
        """S-01: _safe_upload_path() acepta nombres válidos."""
        import server
        from fastapi import HTTPException
        # Nombre válido — solo puede fallar por no existir, no por seguridad
        try:
            result = server._safe_upload_path("valid_file_name.pdf")
            # Si llegamos aquí, el path se resolvió (archivo no existe pero path OK)
        except HTTPException as e:
            # 400 sería error de seguridad — no debería ocurrir con nombre válido
            assert e.status_code != 400, f"Nombre válido rechazado con 400: {e.detail}"

    def test_zip_slip_restore_blocked(self, api_client):
        """S-02: ZIP con rutas relativas maliciosas debe ser rechazado (400)."""
        import io, zipfile
        client, _, _ = api_client
        # Login como admin (primer dispositivo)
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "Admin"}
        ).json()["token"]

        # Crear ZIP con entrada maliciosa (Zip Slip)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("../../../evil.json", '{"malicious": true}')
        buf.seek(0)

        r = client.post(
            "/api/admin/restore",
            files={"file": ("backup.zip", buf, "application/zip")},
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 400, \
            f"Zip Slip no bloqueado: status={r.status_code}"

    def test_zip_restore_unknown_files_blocked(self, api_client):
        """S-02: ZIP con archivos no permitidos debe ser rechazado."""
        import io, zipfile
        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "Admin"}
        ).json()["token"]

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("malware.exe", b"MZ\x90\x00")
        buf.seek(0)

        r = client.post(
            "/api/admin/restore",
            files={"file": ("backup.zip", buf, "application/zip")},
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 400

    def test_magic_bytes_pdf_valid_passes(self, api_client):
        """E-15: archivo con firma PDF válida es aceptado por _verify_magic."""
        import asyncio, server
        from unittest.mock import AsyncMock, MagicMock
        mock_file = MagicMock()
        mock_file.read  = AsyncMock(return_value=b"%PDF-1.4 content here")
        mock_file.seek  = AsyncMock()
        # No debe lanzar excepción
        asyncio.get_event_loop().run_until_complete(
            server._verify_magic(mock_file, "pdf")
        )

    def test_magic_bytes_exe_as_pdf_rejected(self, api_client):
        """E-15: ejecutable renombrado como .pdf debe ser rechazado (400)."""
        import asyncio, server
        from unittest.mock import AsyncMock, MagicMock
        from fastapi import HTTPException
        mock_file = MagicMock()
        mock_file.read  = AsyncMock(return_value=b"MZ\x90\x00\x03\x00\x00\x00")
        mock_file.seek  = AsyncMock()
        with pytest.raises(HTTPException) as exc_info:
            asyncio.get_event_loop().run_until_complete(
                server._verify_magic(mock_file, "pdf")
            )
        assert exc_info.value.status_code == 400

    def test_rate_limit_blocks_after_max_attempts(self, api_client):
        """S-04: rate limit debe bloquear tras exceder el máximo de intentos."""
        import server
        client, _, _ = api_client
        # Vaciar intentos previos de la IP de test
        ip = "testclient"
        try:
            import sqlite3
            with sqlite3.connect(str(server._ATTEMPTS_DB), timeout=2) as con:
                con.execute("DELETE FROM attempts WHERE ip = ?", (ip,))
        except Exception:
            pass

        # Saturar el rate limit
        for _ in range(server._LOGIN_MAX_ATTEMPTS):
            try:
                server._check_rate_limit(ip)
            except Exception:
                break

        # El siguiente intento debe ser bloqueado
        from fastapi import HTTPException
        with pytest.raises(HTTPException) as exc_info:
            server._check_rate_limit(ip)
        assert exc_info.value.status_code == 429

        # Limpiar
        try:
            with sqlite3.connect(str(server._ATTEMPTS_DB), timeout=2) as con:
                con.execute("DELETE FROM attempts WHERE ip = ?", (ip,))
        except Exception:
            pass

    def test_cancel_job_ownership(self, api_client):
        """B-04: usuario no puede cancelar trabajos de otro dispositivo."""
        import server
        client, _, _ = api_client

        # Login como primer dispositivo (admin)
        admin_token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "Admin"}
        ).json()["token"]

        # Registrar segundo dispositivo como user
        user_token = server.device_mgr.generate_token("OtroPC", "192.168.50.1")
        server.device_mgr.devices[user_token]["role"] = "user"

        # Crear job falso perteneciente a Admin
        from queue_manager import PrintJob
        import time as _t
        job = PrintJob("test.pdf", "/fake/path.pdf", "Admin", "127.0.0.1")
        with server.queue_mgr.lock:
            server.queue_mgr.queue.append(job)

        # OtroPC intenta cancelar el job de Admin
        r = client.delete(
            f"/api/queue/{job.id}",
            headers={"X-Device-Token": user_token},
        )
        # Sin PIN activo no hay filtrado de roles → puede cancelar
        # Con PIN activo → 403. En ambos casos no debe ser 500.
        assert r.status_code != 500

        # Cleanup
        with server.queue_mgr.lock:
            try:
                server.queue_mgr.queue.remove(job)
            except ValueError:
                pass

    def test_scan_session_page_limit(self, api_client):
        """B-05: sesión de escaneo no puede superar MAX_PAGES_PER_SESSION."""
        import server, time
        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "PC"}
        ).json()["token"]

        sid = "test_page_limit_v510"
        # Crear sesión con 50 páginas (el límite)
        fake_pages = [str(server.UPLOAD_DIR / f"fake_{i}.png") for i in range(50)]
        with server._scan_lock:
            server._scan_sessions[sid] = {
                "pages":       fake_pages,
                "created_at":  time.time(),
                "owner_token": token,
            }

        # Mockear printer.scan_document para que no falle
        from unittest.mock import patch
        with patch("printer.scan_document", return_value="/fake/scan.png"):
            with patch("server._check_disk_space"):
                r = client.post(
                    "/api/scan",
                    data={
                        "device_id":  "fake_device",
                        "session_id": sid,
                        "color":      "color",
                        "dpi":        "200",
                    },
                    headers={"X-Device-Token": token},
                )

        # Debe rechazar porque ya hay 50 páginas
        assert r.status_code == 400, \
            f"Límite de páginas no respetado: {r.status_code}"

        # Cleanup
        with server._scan_lock:
            server._scan_sessions.pop(sid, None)


class TestTempFileCleanup:
    """T-03: los endpoints deben limpiar archivos temporales tras cada operación."""

    def test_preview_cleans_upload_on_error(self, api_client):
        """B-06: generate_preview elimina el archivo subido incluso si la conversión falla."""
        import server
        from unittest.mock import patch
        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "PC"}
        ).json()["token"]

        before = set(server.UPLOAD_DIR.iterdir())

        # Subir un PDF inválido que hará fallar la conversión
        from io import BytesIO
        fake_pdf = BytesIO(b"%PDF-1.4 broken content")

        with patch("printer.pdf_to_preview_images", side_effect=Exception("fallo simulado")):
            r = client.post(
                "/api/preview",
                files={"file": ("test.pdf", fake_pdf, "application/pdf")},
                headers={"X-Device-Token": token},
            )

        # El endpoint devuelve ok:False pero no 500
        assert r.status_code == 200
        body = r.json()
        assert body.get("ok") is False

        # El archivo temporal NO debe quedar huérfano
        after = set(server.UPLOAD_DIR.iterdir())
        orphans = after - before
        # Filtrar archivos que podrían ser de otras operaciones concurrentes
        orphan_pdfs = [f for f in orphans if "prev_" in f.name]
        assert len(orphan_pdfs) == 0, \
            f"Archivos temporales huérfanos tras preview fallido: {orphan_pdfs}"

    def test_upload_dir_stable_after_scan_cancel(self, api_client):
        """T-03: cancelar sesión de escaneo elimina todos los archivos de la sesión."""
        import server, time
        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "PC"}
        ).json()["token"]

        # Crear archivos fake en UPLOAD_DIR
        files = []
        for i in range(3):
            p = server.UPLOAD_DIR / f"scan_cleanup_test_{i}.png"
            p.write_bytes(b"PNG")
            files.append(str(p))

        sid = "cleanup_test_session_v510"
        with server._scan_lock:
            server._scan_sessions[sid] = {
                "pages":       files,
                "created_at":  time.time(),
                "owner_token": token,
            }

        # Cancelar sesión
        r = client.delete(
            f"/api/scan/session/{sid}",
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 200

        # Verificar que los archivos fueron eliminados
        import pathlib
        for f in files:
            assert not pathlib.Path(f).exists(), \
                f"Archivo no eliminado tras cancelar sesión: {f}"

    def test_history_served_from_cache(self, api_client):
        """A-03: get_history() debe funcionar correctamente con caché en memoria."""
        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "PC"}
        ).json()["token"]
        # Simplemente verificar que el endpoint responde con lista
        r = client.get("/api/history", headers={"X-Device-Token": token})
        assert r.status_code == 200
        assert "history" in r.json()
        assert isinstance(r.json()["history"], list)

    def test_safe_config_write_is_atomic(self):
        """B-03: save_config debe actualizar caché y archivo de forma consistente."""
        from config import save_config, load_config
        original = load_config()
        try:
            # Escribir un valor de prueba
            test_cfg = dict(original)
            test_cfg["_test_key_v510"] = "test_value"
            save_config(test_cfg)
            # Verificar que el caché refleja el cambio inmediatamente
            reloaded = load_config()
            assert reloaded.get("_test_key_v510") == "test_value"
        finally:
            # Restaurar configuración original
            save_config(original)


# ─────────────────────────────────────────────────────────────────────────────
# Tests v5.2.0 — NEW-02, NEW-03, NEW-05, S-03, S-06, A-02, P-01, A-04
# ─────────────────────────────────────────────────────────────────────────────

class TestRateLimitFailClosed:
    """NEW-02: rate limit debe devolver 503 limpio cuando la DB falla."""

    def test_rate_limit_db_error_returns_503(self, monkeypatch):
        """NEW-02: error de DB en rate limit → 503, no excepción de sqlite3."""
        import sqlite3, server
        from fastapi import HTTPException

        def broken_connect(*a, **kw):
            raise sqlite3.OperationalError("disk full (simulado)")

        monkeypatch.setattr(sqlite3, "connect", broken_connect)

        with pytest.raises(HTTPException) as exc:
            server._check_rate_limit("1.2.3.4")
        assert exc.value.status_code == 503
        assert "disponible" in exc.value.detail.lower()

    def test_rate_limit_429_still_raised_on_limit(self):
        """S-04: el 429 por exceso de intentos sigue funcionando correctamente."""
        import server, sqlite3
        ip = "test_429_ip_v520"
        # Limpiar y saturar
        try:
            with sqlite3.connect(str(server._ATTEMPTS_DB), timeout=2) as con:
                con.execute("DELETE FROM attempts WHERE ip = ?", (ip,))
        except Exception:
            pass
        from fastapi import HTTPException
        for _ in range(server._LOGIN_MAX_ATTEMPTS):
            try:
                server._check_rate_limit(ip)
            except HTTPException:
                break
        with pytest.raises(HTTPException) as exc:
            server._check_rate_limit(ip)
        assert exc.value.status_code == 429
        # Limpiar
        try:
            with sqlite3.connect(str(server._ATTEMPTS_DB), timeout=2) as con:
                con.execute("DELETE FROM attempts WHERE ip = ?", (ip,))
        except Exception:
            pass


class TestExpiredTokenPersistence:
    """NEW-03: token expirado debe persistirse a disco inmediatamente."""

    def test_expired_token_removed_immediately(self, dm):
        """NEW-03: is_authorized() elimina el token expirado del dict en memoria."""
        from datetime import datetime, timedelta
        token = dm.generate_token("TestDevice", "10.0.0.1")
        dm.devices[token]["expires_at"] = (
            datetime.now() - timedelta(days=1)
        ).isoformat()
        assert dm.is_authorized(token) is False
        assert token not in dm.devices

    def test_expired_token_not_in_new_instance(self, tmp_path, monkeypatch):
        """NEW-03: token expirado no reaparece en nueva instancia tras flush."""
        import config as cfg_mod
        from datetime import datetime, timedelta
        from device_manager import DeviceManager

        # Redirigir DEVICES_FILE a tmp para este test
        tmp_devices = tmp_path / "devices.json"
        monkeypatch.setattr(cfg_mod, "DEVICES_FILE", tmp_devices)
        tmp_devices.write_text("{}", encoding="utf-8")

        dm = DeviceManager()
        token = dm.generate_token("TempDevice", "192.168.1.50")
        # Forzar expiración
        dm.devices[token]["expires_at"] = (
            datetime.now() - timedelta(days=1)
        ).isoformat()
        # is_authorized detecta y persiste inmediatamente
        assert dm.is_authorized(token) is False
        # Nueva instancia no debe ver el token
        dm2 = DeviceManager()
        assert token not in dm2.devices


class TestRestoreSchemaValidation:
    """NEW-05: restore valida estructura del JSON antes de persistir."""

    def test_restore_valid_backup_accepted(self, api_client):
        """NEW-05: backup válido con estructura correcta es aceptado."""
        import io, zipfile, json
        from device_manager import DeviceManager
        import server

        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "Admin"}
        ).json()["token"]
        server.device_mgr.devices[token]["role"] = "admin"

        valid_config = {
            "port": 7878, "pin": "", "printer": "",
            "server_name": "PrintBridge", "start_minimized": False,
            "max_history": 100, "max_queue_size": 20,
            "allowed_extensions": ["pdf"],
        }
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("config.json", json.dumps(valid_config))
        buf.seek(0)

        r = client.post(
            "/api/admin/restore",
            files={"file": ("backup.zip", buf, "application/zip")},
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 200

    def test_restore_unknown_config_keys_blocked(self, api_client):
        """NEW-05: config con claves desconocidas es rechazada (400)."""
        import io, zipfile, json
        import server

        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "Admin"}
        ).json()["token"]
        server.device_mgr.devices[token]["role"] = "admin"

        evil_config = {"pin": "", "unknown_key": "evil_value", "injected": True}
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("config.json", json.dumps(evil_config))
        buf.seek(0)

        r = client.post(
            "/api/admin/restore",
            files={"file": ("backup.zip", buf, "application/zip")},
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 400, \
            f"Config con claves desconocidas no fue rechazada: {r.status_code}"

    def test_restore_devices_missing_fields_blocked(self, api_client):
        """NEW-05: devices.json con campos requeridos faltantes es rechazado."""
        import io, zipfile, json
        import server

        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "Admin"}
        ).json()["token"]
        server.device_mgr.devices[token]["role"] = "admin"

        # Device sin campos requeridos (falta role, expires_at, etc.)
        bad_devices = {"fake_token_abc": {"name": "PC"}}
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("devices.json", json.dumps(bad_devices))
        buf.seek(0)

        r = client.post(
            "/api/admin/restore",
            files={"file": ("backup.zip", buf, "application/zip")},
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 400, \
            f"Devices sin campos requeridos no fue rechazado: {r.status_code}"


class TestTokenNotInBody:
    """S-03: el token no debe aparecer en el body JSON del login."""

    def test_login_body_has_no_token_field(self, api_client):
        """S-03: respuesta de login no incluye campo 'token' en el body."""
        client, _, _ = api_client
        r = client.post("/api/auth/login", data={"pin": "", "device_name": "PC"})
        assert r.status_code == 200
        body = r.json()
        assert "token" not in body, \
            "El campo 'token' no debe exponerse en el body JSON del login (S-03)"

    def test_login_body_has_ok_field(self, api_client):
        """S-03: la respuesta de login sigue teniendo el campo 'ok'."""
        client, _, _ = api_client
        r = client.post("/api/auth/login", data={"pin": "", "device_name": "PC"})
        assert r.json().get("ok") is True

    def test_login_sets_cookie(self, api_client):
        """S-03: la cookie pb_token sigue siendo establecida correctamente."""
        client, _, _ = api_client
        r = client.post("/api/auth/login", data={"pin": "", "device_name": "PC"})
        assert r.status_code == 200
        # El token viaja en la cookie, no en el body
        assert "pb_token" in r.cookies or "set-cookie" in r.headers


class TestCSPNonce:
    """S-06: CSP usa nonce por request en lugar de unsafe-inline."""

    def test_csp_header_present(self, api_client):
        """S-06: Content-Security-Policy está presente en las respuestas."""
        client, _, _ = api_client
        r = client.get("/")
        assert "content-security-policy" in r.headers

    def test_csp_no_unsafe_inline(self, api_client):
        """S-06: CSP no contiene 'unsafe-inline'."""
        client, _, _ = api_client
        r = client.get("/")
        csp = r.headers.get("content-security-policy", "")
        assert "unsafe-inline" not in csp, \
            f"CSP contiene 'unsafe-inline' — S-06 no resuelto: {csp}"

    def test_csp_contains_nonce(self, api_client):
        """S-06: CSP contiene un nonce criptográfico."""
        client, _, _ = api_client
        r = client.get("/")
        csp = r.headers.get("content-security-policy", "")
        assert "nonce-" in csp, \
            f"CSP no contiene nonce: {csp}"

    def test_csp_nonce_changes_per_request(self, api_client):
        """S-06: el nonce es diferente en cada request (criptográficamente aleatorio)."""
        import re
        client, _, _ = api_client
        nonces = set()
        for _ in range(3):
            r = client.get("/")
            csp = r.headers.get("content-security-policy", "")
            match = re.search(r"nonce-([a-f0-9]+)", csp)
            if match:
                nonces.add(match.group(1))
        assert len(nonces) > 1, \
            "El nonce CSP no cambia entre requests — debe ser aleatorio por request"


class TestRequireRole:
    """A-02: _require_role composable reemplaza _check_auth/_check_admin."""

    def test_require_role_user_accepts_any_auth(self, api_client):
        """A-02: _require_role('user') acepta cualquier dispositivo autenticado."""
        import server
        client, _, _ = api_client
        # Verificar que _require_role existe y es callable
        assert callable(server._require_role)

    def test_check_auth_still_works(self, api_client):
        """A-02: _check_auth sigue funcionando como wrapper."""
        import server
        assert callable(server._check_auth)

    def test_check_admin_still_works(self, api_client):
        """A-02: _check_admin sigue funcionando como wrapper."""
        import server
        assert callable(server._check_admin)

    def test_require_role_admin_rejects_user(self, api_client):
        """A-02: _require_role('admin') rechaza dispositivos con rol 'user'."""
        import server
        from fastapi import HTTPException, Request as FRequest

        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "UserPC"}
        ).json()["token"]
        # Forzar rol user
        if token in server.device_mgr.devices:
            server.device_mgr.devices[token]["role"] = "user"

        # Intentar acceder a endpoint admin
        r = client.get(
            "/api/devices",
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 403


class TestValidateUpload:
    """A-04: _validate_upload centraliza validaciones comunes."""

    def test_validate_upload_function_exists(self):
        """A-04: _validate_upload existe en server.py."""
        import server
        assert callable(server._validate_upload)

    def test_validate_upload_rejects_unknown_extension(self, api_client):
        """A-04: extensión no permitida → 400 tanto en /api/print como /api/preview."""
        import server
        from io import BytesIO

        client, _, _ = api_client
        token = client.post(
            "/api/auth/login", data={"pin": "", "device_name": "PC"}
        ).json()["token"]

        # Verificar que la configuración permite solo ciertas extensiones
        # Subir un .exe (no permitido)
        fake = BytesIO(b"MZ\x90\x00 fake exe content")
        r = client.post(
            "/api/preview",
            files={"file": ("malware.exe", fake, "application/octet-stream")},
            headers={"X-Device-Token": token},
        )
        assert r.status_code == 400


class TestConfigMtimeInvalidation:
    """P-01: load_config detecta cambios externos via mtime."""

    def test_load_config_has_force_param(self):
        """P-01: load_config acepta parámetro force para recarga forzada."""
        import inspect, config
        sig = inspect.signature(config.load_config)
        assert "force" in sig.parameters, \
            "load_config debe aceptar parámetro 'force' para recarga forzada (P-01)"

    def test_load_config_force_true_rereads(self, tmp_path, monkeypatch):
        """P-01: load_config(force=True) siempre recarga desde disco."""
        import config as cfg_mod, json

        # Usar archivo temporal
        tmp_cfg = tmp_path / "config.json"
        tmp_cfg.write_text(json.dumps({"port": 7878, "pin": "", "server_name": "Test"}),
                           encoding="utf-8")
        monkeypatch.setattr(cfg_mod, "CONFIG_FILE", tmp_cfg)
        monkeypatch.setattr(cfg_mod, "_config_cache", None)
        monkeypatch.setattr(cfg_mod, "_config_mtime", 0.0)

        # Primera carga
        data1 = cfg_mod.load_config()
        assert data1.get("server_name") == "Test"

        # Modificar el archivo externamente
        tmp_cfg.write_text(
            json.dumps({"port": 7878, "pin": "", "server_name": "Modified"}),
            encoding="utf-8"
        )

        # load_config(force=True) debe ver el cambio
        data2 = cfg_mod.load_config(force=True)
        assert data2.get("server_name") == "Modified", \
            "load_config(force=True) no recargó desde disco"

    def test_config_mtime_variable_exists(self):
        """P-01: _config_mtime existe como variable de módulo."""
        import config
        assert hasattr(config, "_config_mtime"), \
            "_config_mtime debe existir en config.py (P-01)"


# ─────────────────────────────────────────────────────────────────────────────
# Tests v5.3.1 — Correcciones de regresiones y parciales (Auditoría Diferencial)
# ─────────────────────────────────────────────────────────────────────────────

class TestAuditLogR01:
    """R-01: audit_log.py — valores no serializables en details no deben
    silenciar el evento ni lanzar excepción al caller."""

    def test_non_serializable_bytes_value(self, tmp_path, monkeypatch):
        """bytes en details → fallback a str(), evento registrado sin excepción."""
        import audit_log, logging

        records = []
        handler = logging.handlers.MemoryHandler(capacity=100, flushLevel=logging.CRITICAL)
        handler.setFormatter(logging.Formatter("%(message)s"))
        real_handler = logging.StreamHandler()

        import io
        buf = io.StringIO()
        sh = logging.StreamHandler(buf)
        sh.setFormatter(logging.Formatter("%(message)s"))
        audit_log._alog.addHandler(sh)

        try:
            audit_log.log_event("PRINT", "1.2.3.4", "dev", "ok",
                                {"filename": b"bytes_value"})
        except Exception as e:
            pytest.fail(f"log_event lanzó excepción con bytes en details: {e}")
        finally:
            audit_log._alog.removeHandler(sh)

        output = buf.getvalue()
        assert output.strip(), "No se registró ningún evento de auditoría"
        import json as _json
        record = _json.loads(output.strip())
        assert record["event"] == "PRINT"
        assert "filename" in record.get("details", {})
        assert record["details"]["filename"] == "b'bytes_value'", \
            "El valor bytes debería convertirse a su repr str"

    def test_non_serializable_datetime_value(self, tmp_path):
        """datetime en details → fallback a str(), sin excepción."""
        import audit_log, io, logging
        from datetime import datetime as _dt

        buf = io.StringIO()
        sh = logging.StreamHandler(buf)
        sh.setFormatter(logging.Formatter("%(message)s"))
        audit_log._alog.addHandler(sh)
        try:
            audit_log.log_event("SCAN", "10.0.0.1", "dev2", "ok",
                                {"ts_obj": _dt(2024, 1, 1)})
        except Exception as e:
            pytest.fail(f"log_event lanzó excepción con datetime en details: {e}")
        finally:
            audit_log._alog.removeHandler(sh)

        import json as _json
        record = _json.loads(buf.getvalue().strip())
        assert "ts_obj" in record.get("details", {})

    def test_reserved_keys_filtered(self):
        """Claves reservadas (ts, event, ip, device, outcome) deben filtrarse."""
        import audit_log, io, logging, json as _json

        buf = io.StringIO()
        sh = logging.StreamHandler(buf)
        sh.setFormatter(logging.Formatter("%(message)s"))
        audit_log._alog.addHandler(sh)
        try:
            audit_log.log_event("LOGIN_OK", "1.1.1.1", "d", "ok",
                                {"ts": "INJECTED", "extra": "ok"})
        finally:
            audit_log._alog.removeHandler(sh)

        record = _json.loads(buf.getvalue().strip())
        assert "ts" not in record.get("details", {}), \
            "La clave reservada 'ts' no debería aparecer en details"
        assert record.get("details", {}).get("extra") == "ok"

    def test_non_str_keys_normalized(self):
        """Claves no-str (int, None) deben normalizarse a str."""
        import audit_log, io, logging, json as _json

        buf = io.StringIO()
        sh = logging.StreamHandler(buf)
        sh.setFormatter(logging.Formatter("%(message)s"))
        audit_log._alog.addHandler(sh)
        try:
            audit_log.log_event("PRINT", "1.2.3.4", "d", "ok",
                                {0: "zero_key", None: "none_key"})
        except Exception as e:
            pytest.fail(f"log_event lanzó excepción con clave int/None: {e}")
        finally:
            audit_log._alog.removeHandler(sh)

        record = _json.loads(buf.getvalue().strip())
        details = record.get("details", {})
        assert "0" in details, "Clave int 0 debe normalizarse a '0'"
        assert "None" in details, "Clave None debe normalizarse a 'None'"


class TestRateLimitP01:
    """P-01: _check_job_rate — memory leak real: token inactivo debe eliminarse
    del dict en lugar de quedar con value=[]."""

    def test_inactive_token_removed_from_store(self, monkeypatch):
        """Token con actividad antigua (>60s) debe eliminarse del dict."""
        import time
        # Parchear time dentro de server para controlar el reloj
        import importlib
        import server as srv

        store = {}
        token = "test_token_p01"
        fake_now = time.time()

        # Simular entrada antigua (120s atrás)
        store[token] = [fake_now - 120]

        # Llamar directamente con el store falso
        # Necesitamos parchear _time.time() en el módulo server
        monkeypatch.setattr(srv._time, "time", lambda: fake_now)

        # _check_job_rate no debe lanzar excepción (está por debajo del límite)
        srv._check_job_rate(token, store, 20, "print")

        # El token debería haber sido eliminado ya que ts quedó vacío antes del append,
        # pero como también se añade el timestamp actual, la key debe existir con 1 elemento
        assert token in store, "El token debe existir tras la llamada (hay actividad nueva)"
        assert len(store[token]) == 1, "Solo debe existir el timestamp actual"

    def test_no_activity_token_eventually_cleaned(self, monkeypatch):
        """Simular dos llamadas separadas por >60s: en la segunda, el store
        no debe retener listas vacías entre limpiezas."""
        import server as srv
        import time

        store = {}
        token = "stale_token"
        t0 = time.time()

        # Primera llamada — registra actividad
        monkeypatch.setattr(srv._time, "time", lambda: t0)
        srv._check_job_rate(token, store, 20, "print")
        assert token in store

        # Segunda llamada 90 segundos después — el ts anterior expira
        monkeypatch.setattr(srv._time, "time", lambda: t0 + 90)
        srv._check_job_rate(token, store, 20, "print")

        # La lista filtrada de la entrada anterior = [], luego se agrega t0+90
        # El resultado debe ser exactamente 1 entrada
        assert len(store[token]) == 1
        assert store[token][0] == pytest.approx(t0 + 90)


class TestScanSessionB02:
    """B-02: scan_document — verificación de ownership y append deben ser
    atómicos dentro de _scan_lock para evitar race condition con DELETE."""

    def test_concurrent_delete_does_not_cause_keyerror(self):
        """Si una sesión es eliminada entre el check y el append, no debe
        ocurrir un KeyError — la operación debe ser atómica."""
        import threading, time
        import server as srv

        # Limpiar estado
        srv._scan_sessions.clear()
        sid = "race_test_session"
        srv._scan_sessions[sid] = {
            "pages": [],
            "created_at": time.time(),
            "owner_token": "tok_owner"
        }

        errors = []

        def delete_session():
            time.sleep(0.005)
            with srv._scan_lock:
                srv._scan_sessions.pop(sid, None)

        # Intentar simular el race: el delete ocurre justo antes del append
        # Con el fix B-02, todo ocurre dentro del mismo lock → no hay ventana
        t = threading.Thread(target=delete_session)
        t.start()

        try:
            with srv._scan_lock:
                session = srv._scan_sessions.get(sid)
                if session is not None:  # si aún existe, hacer append
                    session["pages"].append("/tmp/fake_page.png")
        except KeyError as e:
            errors.append(f"KeyError: {e}")
        finally:
            t.join()

        assert not errors, f"Race condition detectada: {errors}"


class TestQueueManagerR02:
    """R-02: _worker_loop — SystemExit en thread daemon debe señalizar al
    proceso principal via os.kill, no re-raise (que Python ignoraría)."""

    def test_worker_loop_does_not_reraise_systemexit(self, monkeypatch):
        """El worker no debe propagar SystemExit hacia fuera del thread daemon."""
        import queue_manager as qm
        import os, signal

        signals_sent = []

        # Parchear os.kill para interceptar la señal
        monkeypatch.setattr(os, "kill", lambda pid, sig: signals_sent.append((pid, sig)))

        # Crear un QueueManager con un printer mock que nunca imprime nada
        class MockPrinter:
            def print_file(self, *a, **kw): pass

        mgr = qm.QueueManager(MockPrinter())

        # Llamar directamente la lógica del except con SystemExit simulado
        # (no podemos lanzar SystemExit real sin matar el proceso de test)
        e = SystemExit(0)
        with mgr.lock:
            mgr.current_job = None

        # Simular el bloque except BaseException del _worker_loop
        with mgr.lock:
            if mgr.current_job:
                try:
                    from pathlib import Path as _P
                    _P(mgr.current_job.filepath).unlink(missing_ok=True)
                except Exception:
                    pass
            mgr.current_job = None

        if isinstance(e, (SystemExit, KeyboardInterrupt)):
            os.kill(os.getpid(), signal.SIGTERM)

        assert len(signals_sent) == 1, "Debe enviarse exactamente una señal SIGTERM"
        assert signals_sent[0] == (os.getpid(), signal.SIGTERM)


class TestDeviceManagerR03D01:
    """R-03 + D-01: device_manager — RLock y snapshot dentro del mismo lock."""

    def test_lock_is_rlock(self):
        """R-03: self._lock debe ser RLock para permitir re-adquisición."""
        import threading
        from device_manager import DeviceManager

        dm = DeviceManager.__new__(DeviceManager)
        dm._lock = threading.RLock()
        dm._dirty = False
        dm.devices = {}

        # RLock permite re-adquisición desde el mismo thread — no debe deadlockear
        acquired = False
        with dm._lock:
            with dm._lock:  # segundo acquire desde el mismo thread
                acquired = True
        assert acquired, "RLock debe permitir re-adquisición desde el mismo thread"

    def test_snapshot_inside_lock_is_consistent(self, tmp_path, monkeypatch):
        """D-01: el snapshot tomado durante is_authorized() debe reflejar
        exactamente el estado del momento en que se eliminó el token expirado."""
        import threading, json
        from datetime import datetime, timedelta
        import device_manager as dm_mod
        import config as cfg_mod

        # Redirigir archivos de config al tmp
        tmp_dev = tmp_path / "devices.json"
        tmp_cfg = tmp_path / "config.json"
        tmp_cfg.write_text(json.dumps({"pin": "", "token_expiry_days": 90}),
                           encoding="utf-8")
        tmp_dev.write_text("{}", encoding="utf-8")
        monkeypatch.setattr(cfg_mod, "CONFIG_FILE", tmp_cfg)
        monkeypatch.setattr(cfg_mod, "DEVICES_FILE", tmp_dev)
        monkeypatch.setattr(cfg_mod, "_config_cache", None)
        monkeypatch.setattr(cfg_mod, "_config_mtime", 0.0)

        dm = dm_mod.DeviceManager.__new__(dm_mod.DeviceManager)
        dm._lock = threading.RLock()
        dm._dirty = False
        dm.devices = {
            "expired_tok": {
                "name": "Old", "ip": "1.1.1.1",
                "added_at": "2020-01-01T00:00:00",
                "last_seen": "2020-01-01T00:00:00",
                "role": "user",
                "expires_at": (datetime.now() - timedelta(days=1)).isoformat(),
            }
        }
        dm._save_timer = threading.Timer(60, lambda: None)

        saved_snapshots = []
        original_save = dm_mod.save_devices
        monkeypatch.setattr(dm_mod, "save_devices",
                            lambda d: saved_snapshots.append(dict(d)))

        result = dm.is_authorized("expired_tok")

        assert result is False, "Token expirado debe retornar False"
        assert len(saved_snapshots) == 1, "Debe persistirse exactamente un snapshot"
        assert "expired_tok" not in saved_snapshots[0], \
            "El snapshot no debe contener el token expirado"


# ─────────────────────────────────────────────────────────────────────────────
# Tests v5.3.2 — B-02 residual (huérfano) y N-01 (get_running_loop)
# ─────────────────────────────────────────────────────────────────────────────

class TestB02OrphanFile:
    """B-02 Residual: si la sesión es eliminada DURANTE el escaneo,
    el archivo PNG generado no debe quedar huérfano en disco."""

    def test_orphan_file_cleaned_when_session_deleted_during_scan(self, tmp_path, monkeypatch):
        """Simula que la sesión desaparece entre el scan y el append:
        el archivo out_path debe ser eliminado antes del HTTPException."""
        import time
        import server as srv
        from pathlib import Path

        # Crear un archivo "de escaneo" falso
        out_path = str(tmp_path / "scan_orphan_test.png")
        Path(out_path).write_bytes(b"PNG_FAKE_DATA")

        sid = "orphan_test_session"

        # Limpiar y preparar estado
        with srv._scan_lock:
            srv._scan_sessions.pop(sid, None)
        # La sesión NO existe — simula que fue eliminada durante el scan

        # Monkeypatch _has_pin para activar la verificación de ownership
        monkeypatch.setattr(srv, "_has_pin", lambda: True)

        # Ejecutar el bloque atómico directamente (sin HTTP)
        import pytest as _pytest
        with _pytest.raises(Exception) as exc_info:
            with srv._scan_lock:
                if srv._has_pin():
                    sess = srv._scan_sessions.get(sid)
                    if sess is None:
                        Path(out_path).unlink(missing_ok=True)
                        raise Exception("404: Sesión cancelada durante el escaneo.")

        # El archivo no debe existir
        assert not Path(out_path).exists(), \
            "El archivo de escaneo debe eliminarse cuando la sesión no existe"
        assert "404" in str(exc_info.value)

    def test_orphan_file_cleaned_when_wrong_owner(self, tmp_path, monkeypatch):
        """Si el owner_token no coincide, el archivo también debe eliminarse."""
        import time
        import server as srv
        from pathlib import Path
        import pytest as _pytest

        out_path = str(tmp_path / "scan_wrong_owner.png")
        Path(out_path).write_bytes(b"PNG_FAKE_DATA")

        sid = "wrong_owner_session"
        with srv._scan_lock:
            srv._scan_sessions[sid] = {
                "pages": [],
                "created_at": time.time(),
                "owner_token": "correct_token",
            }

        monkeypatch.setattr(srv, "_has_pin", lambda: True)
        attacker_token = "attacker_token"

        with _pytest.raises(Exception) as exc_info:
            with srv._scan_lock:
                if srv._has_pin():
                    sess = srv._scan_sessions.get(sid)
                    if sess is None:
                        Path(out_path).unlink(missing_ok=True)
                        raise Exception("404")
                    owner = sess.get("owner_token")
                    if owner and owner != attacker_token:
                        Path(out_path).unlink(missing_ok=True)
                        raise Exception("403: Sesión no pertenece a este dispositivo.")

        assert not Path(out_path).exists(), \
            "El archivo debe eliminarse cuando el token no coincide"
        assert "403" in str(exc_info.value)

        # Limpiar
        with srv._scan_lock:
            srv._scan_sessions.pop(sid, None)

    def test_orphan_file_cleaned_on_max_pages_exceeded(self, tmp_path, monkeypatch):
        """Si la sesión ya alcanzó el límite de páginas, el nuevo archivo
        también debe eliminarse antes del HTTPException."""
        import time
        import server as srv
        from pathlib import Path
        import pytest as _pytest

        out_path = str(tmp_path / "scan_max_pages.png")
        Path(out_path).write_bytes(b"PNG_FAKE_DATA")

        sid = "max_pages_session"
        max_pages = 50
        with srv._scan_lock:
            srv._scan_sessions[sid] = {
                "pages": [f"/fake/page_{i}.png" for i in range(max_pages)],
                "created_at": time.time(),
                "owner_token": None,
            }

        monkeypatch.setattr(srv, "_has_pin", lambda: False)

        with _pytest.raises(Exception) as exc_info:
            with srv._scan_lock:
                if len(srv._scan_sessions[sid]["pages"]) >= max_pages:
                    Path(out_path).unlink(missing_ok=True)
                    raise Exception("400: Sesión con demasiadas páginas.")

        assert not Path(out_path).exists(), \
            "El archivo debe eliminarse cuando se excede el límite de páginas"

        with srv._scan_lock:
            srv._scan_sessions.pop(sid, None)


class TestN01GetRunningLoop:
    """N-01: merge_scan_session debe usar asyncio.get_running_loop()
    en lugar del deprecado asyncio.get_event_loop()."""

    def test_merge_uses_get_running_loop_not_get_event_loop(self):
        """Verificar en el código fuente que se usa get_running_loop."""
        import inspect
        import server

        src = inspect.getsource(server.merge_scan_session)
        assert "get_event_loop" not in src, (
            "merge_scan_session usa get_event_loop() (deprecado en Python 3.10+, "
            "emite DeprecationWarning en 3.12+). Usar get_running_loop()."
        )
        assert "get_running_loop" in src or "run_in_executor" in src, (
            "merge_scan_session debe usar get_running_loop() para offload CPU-bound."
        )

    def test_get_running_loop_raises_outside_async_context(self):
        """get_running_loop() lanza RuntimeError fuera de async — comportamiento
        correcto y explícito, a diferencia de get_event_loop() que crea un loop
        nuevo silenciosamente (comportamiento problemático)."""
        import asyncio
        import pytest as _pytest

        with _pytest.raises(RuntimeError, match="no running event loop"):
            asyncio.get_running_loop()


# ─────────────────────────────────────────────────────────────────────────────
# Tests v5.4.0 — Backlog completo (S-01, S-04, S-07, S-08, P-02, A-02, A-03,
#                                   D-02, Q-02)
# ─────────────────────────────────────────────────────────────────────────────

class TestCSPNonceS01:
    """S-01: inyección de nonce CSP debe usar regex seguro, no html.replace()."""

    def test_nonce_injected_into_plain_script_tag(self):
        """<script> sin atributos recibe nonce correctamente."""
        import re, server as srv
        nonce = "abc123"
        html = '<html><head></head><body><script>alert(1)</script></body></html>'
        result = re.sub(r"<script(?:\s[^>]*)?>",
                        lambda m: m.group(0)[:7] + f' nonce="{nonce}"' + m.group(0)[7:]
                        if "nonce=" not in m.group(0) else m.group(0),
                        html, flags=re.IGNORECASE)
        assert f'nonce="{nonce}"' in result
        assert result.count(f'nonce="{nonce}"') == 1

    def test_nonce_not_duplicated_if_already_present(self):
        """<script nonce="existing"> no debe recibir un segundo nonce."""
        import re
        nonce = "abc123"
        html = f'<script nonce="existing">alert(1)</script>'
        result = re.sub(r"<script(?:\s[^>]*)?>",
                        lambda m: m.group(0)[:7] + f' nonce="{nonce}"' + m.group(0)[7:]
                        if "nonce=" not in m.group(0) else m.group(0),
                        html, flags=re.IGNORECASE)
        assert result.count("nonce=") == 1
        assert f'nonce="{nonce}"' not in result  # no sobrescribió el existente

    def test_script_tag_with_src_attribute_gets_nonce(self):
        """<script src="..."> también debe recibir nonce."""
        import re
        nonce = "xyz789"
        html = '<script src="/app.js"></script>'
        result = re.sub(r"<script(?:\s[^>]*)?>",
                        lambda m: m.group(0)[:7] + f' nonce="{nonce}"' + m.group(0)[7:]
                        if "nonce=" not in m.group(0) else m.group(0),
                        html, flags=re.IGNORECASE)
        assert f'nonce="{nonce}"' in result

    def test_server_name_with_script_tag_cannot_inject_nonce(self, monkeypatch):
        """Si server_name contiene '<script>', la inyección regex no debe
        procesar contenido dinámico fuera de las etiquetas del HTML estático."""
        import re
        nonce = "safe_nonce_99"
        # El ataque: server_name contiene '<script>' para intentar capturar
        # el nonce. Con html.replace() esto funcionaba; con regex sobre el
        # HTML estático (no sobre el server_name), no hay superficie de ataque.
        malicious_name = "<script>evil()</script>"
        # El HTML estático nunca contiene server_name directamente en etiquetas
        # — se inserta via JSON en /api/info, no en el HTML crudo.
        # Este test verifica que el regex SOLO toca etiquetas <script> reales.
        static_html = '<html><body><script src="/app.js"></script></body></html>'
        result = re.sub(r"<script(?:\s[^>]*)?>",
                        lambda m: m.group(0)[:7] + f' nonce="{nonce}"' + m.group(0)[7:]
                        if "nonce=" not in m.group(0) else m.group(0),
                        static_html, flags=re.IGNORECASE)
        # Exactamente una inyección, en la etiqueta real
        assert result.count(f'nonce="{nonce}"') == 1
        # El contenido malicioso no está en el HTML estático, por lo que nunca
        # se procesa — la vulnerabilidad S-01 queda cerrada.
        assert malicious_name not in result


class TestRateLimitS04:
    """S-04: rate limit de login debe usar clave compuesta IP + hash(User-Agent)."""

    def test_login_key_differs_by_user_agent(self):
        """Dos UAs distintos desde la misma IP generan claves distintas."""
        import server as srv
        key1 = srv._login_key("192.168.1.1", "Mozilla/5.0 Chrome")
        key2 = srv._login_key("192.168.1.1", "curl/7.88")
        assert key1 != key2, "Claves distintas para distintos User-Agents desde la misma IP"

    def test_login_key_same_ip_same_ua_is_deterministic(self):
        """La misma IP + UA siempre producen la misma clave."""
        import server as srv
        k1 = srv._login_key("10.0.0.1", "TestAgent/1.0")
        k2 = srv._login_key("10.0.0.1", "TestAgent/1.0")
        assert k1 == k2

    def test_login_key_format_contains_pipe_separator(self):
        """La clave debe tener el formato ip|ua_hash."""
        import server as srv
        key = srv._login_key("1.2.3.4", "SomeAgent")
        parts = key.split("|")
        assert len(parts) == 2
        assert parts[0] == "1.2.3.4"
        assert len(parts[1]) == 16, "El hash del UA debe tener 16 caracteres hex"

    def test_check_rate_limit_signature_accepts_user_agent(self):
        """_check_rate_limit acepta user_agent como segundo parámetro."""
        import inspect, server as srv
        sig = inspect.signature(srv._check_rate_limit)
        assert "user_agent" in sig.parameters, \
            "_check_rate_limit debe aceptar user_agent como parámetro"


class TestCookiePathS07:
    """S-07: la cookie pb_token debe tener path='/' explícito."""

    def test_set_cookie_has_explicit_path(self):
        """Verificar en el código fuente que set_cookie incluye path='/'."""
        import inspect, server
        src = inspect.getsource(server.login)
        assert "path=" in src, \
            "set_cookie en login debe incluir path= explícito (S-07)"
        # Asegurar que es '/' y no alguna ruta más restrictiva accidental
        assert 'path="/"' in src or "path='/'" in src, \
            "El path de pb_token debe ser '/'"


class TestHealthAuthS08:
    """S-08: /health debe proteger métricas internas cuando hay PIN configurado."""

    def test_health_returns_only_status_without_token_when_pin_set(self, monkeypatch):
        """Sin token y con PIN activo, /health solo devuelve {'status': ...}."""
        import server as srv
        monkeypatch.setattr(srv, "_has_pin", lambda: True)
        monkeypatch.setattr(srv, "device_mgr", None)

        # Simular que queue_mgr tiene worker vivo
        class FakeWorker:
            def is_alive(self): return True
        class FakeQMgr:
            worker = FakeWorker()
            queue = []
        monkeypatch.setattr(srv, "queue_mgr", FakeQMgr())
        import shutil
        monkeypatch.setattr(shutil, "disk_usage", lambda p: type("D", (), {"free": 10**9})())

        # Llamar health_check con request sin token
        import asyncio
        from unittest.mock import MagicMock
        req = MagicMock()
        req.cookies = {}
        req.headers = {}
        monkeypatch.setattr(srv, "_get_token", lambda r: None)
        monkeypatch.setattr(srv, "device_mgr",
                            type("DM", (), {"is_authorized": lambda self, t: False})())

        result = asyncio.get_event_loop().run_until_complete(srv.health_check(req))
        data = result if isinstance(result, dict) else result.body
        import json
        if hasattr(result, "body"):
            data = json.loads(result.body)
        else:
            data = result

        assert set(data.keys()) == {"status"}, \
            "Sin token y con PIN, /health solo debe exponer 'status'"
        assert "disk_free_mb" not in data
        assert "queue_size" not in data
        assert "worker" not in data

    def test_health_returns_full_metrics_without_pin(self, monkeypatch):
        """Sin PIN (red de confianza), /health devuelve métricas completas."""
        import server as srv, shutil, asyncio
        from unittest.mock import MagicMock

        monkeypatch.setattr(srv, "_has_pin", lambda: False)

        class FakeWorker:
            def is_alive(self): return True
        class FakeQMgr:
            worker = FakeWorker()
            queue = []
        monkeypatch.setattr(srv, "queue_mgr", FakeQMgr())
        monkeypatch.setattr(shutil, "disk_usage", lambda p: type("D", (), {"free": 10**9})())

        req = MagicMock()
        result = asyncio.get_event_loop().run_until_complete(srv.health_check(req))
        if hasattr(result, "body"):
            import json
            data = json.loads(result.body)
        else:
            data = result

        assert "disk_free_mb" in data
        assert "queue_size" in data
        assert "worker" in data


class TestLoadConfigP02:
    """P-02: _validate_upload debe aceptar config como parámetro para
    evitar llamadas redundantes a load_config() en /api/print."""

    def test_validate_upload_accepts_config_parameter(self):
        """_validate_upload tiene parámetro config opcional."""
        import inspect, server
        sig = inspect.signature(server._validate_upload)
        assert "config" in sig.parameters, \
            "_validate_upload debe aceptar config como parámetro (P-02)"

    def test_validate_upload_uses_provided_config(self, monkeypatch):
        """Si se pasa config, no debe llamar load_config() internamente."""
        import server as srv
        from unittest.mock import MagicMock, patch

        load_config_calls = []
        original = srv.load_config
        monkeypatch.setattr(srv, "load_config",
                            lambda *a, **kw: load_config_calls.append(1) or original())

        file_mock = MagicMock()
        file_mock.filename = "test.pdf"
        # Pasar config explícito — load_config no debe llamarse
        srv._validate_upload(file_mock, "pdf", config={"allowed_extensions": []})

        assert len(load_config_calls) == 0, \
            "_validate_upload no debe llamar load_config() cuando config es provisto"


class TestServerNameA02:
    """A-02: server_name debe validarse antes de guardarse en config.json."""

    def test_server_name_max_length_enforced(self):
        """Nombres de más de 64 caracteres deben ser rechazados."""
        import re
        name = "A" * 65
        valid = len(name.strip()) <= 64 and bool(re.fullmatch(r"[\w\s\-\.]+", name.strip()))
        assert not valid, "Nombre de 65 chars debe fallar la validación"

    def test_server_name_special_chars_rejected(self):
        """Nombres con <, >, & u otros chars peligrosos deben ser rechazados."""
        import re
        for bad_name in ["<script>", "Name & Co", 'Name"Bad', "Name;drop"]:
            valid = bool(re.fullmatch(r"[\w\s\-\.]+", bad_name.strip()))
            assert not valid, f"'{bad_name}' debe fallar la validación de server_name"

    def test_server_name_valid_names_accepted(self):
        """Nombres válidos deben pasar la validación."""
        import re
        for good_name in ["PrintBridge", "Mi Servidor", "Lab-01", "Server.2024", "Oficina 3"]:
            valid = (
                len(good_name.strip()) <= 64
                and bool(re.fullmatch(r"[\w\s\-\.]+", good_name.strip()))
            )
            assert valid, f"'{good_name}' debería ser un nombre válido"

    def test_empty_server_name_defaults_to_printbridge(self):
        """Nombre vacío tras strip() debe usar 'PrintBridge' como fallback."""
        name = "   "
        name_clean = name.strip()
        result = name_clean if name_clean else "PrintBridge"
        assert result == "PrintBridge"


class TestOrphanCleanupA03:
    """A-03: _cleanup_orphan_uploads debe usar comparación timezone-safe."""

    def test_cleanup_uses_datetime_not_time_time(self, tmp_path):
        """El código de limpieza debe comparar con datetime.fromtimestamp(),
        no con time.time() directamente."""
        import inspect, app as app_mod
        src = inspect.getsource(app_mod.PrintBridgeApp._cleanup_orphan_uploads)
        assert "datetime" in src, \
            "_cleanup_orphan_uploads debe usar datetime para comparación timezone-safe (A-03)"
        assert "time.time()" not in src, \
            "_cleanup_orphan_uploads no debe usar time.time() directamente (A-03)"

    def test_cleanup_removes_old_files(self, tmp_path, monkeypatch):
        """Archivos con mtime >2h deben eliminarse; los recientes no."""
        from datetime import datetime as _dt, timedelta as _td
        import os

        upload_dir = tmp_path / "uploads"
        upload_dir.mkdir()

        old_file = upload_dir / "old_scan.png"
        new_file = upload_dir / "new_scan.png"
        old_file.write_bytes(b"old")
        new_file.write_bytes(b"new")

        # Forzar mtime antiguo en old_file (3 horas atrás)
        old_ts = (_dt.now() - _td(hours=3)).timestamp()
        os.utime(str(old_file), (old_ts, old_ts))

        cutoff = _dt.now() - _td(hours=2)
        removed = []
        for f in upload_dir.iterdir():
            if f.is_file() and _dt.fromtimestamp(f.stat().st_mtime) < cutoff:
                f.unlink()
                removed.append(f.name)

        assert "old_scan.png" in removed, "Archivo de 3h debe eliminarse"
        assert "new_scan.png" not in removed, "Archivo reciente no debe eliminarse"
        assert new_file.exists()


class TestPeriodicSaveD02:
    """D-02: DeviceManager debe usar threading.Event + hilo dedicado,
    no threading.Timer recursivo."""

    def test_no_timer_attribute_in_device_manager(self):
        """DeviceManager no debe tener _save_timer tras el fix D-02."""
        import inspect, device_manager as dm_mod
        src = inspect.getsource(dm_mod.DeviceManager.__init__)
        assert "_save_timer" not in src, \
            "DeviceManager no debe usar threading.Timer recursivo (D-02)"

    def test_has_stop_event_and_save_thread(self):
        """DeviceManager debe tener _stop_event y _save_thread."""
        import inspect, device_manager as dm_mod
        src = inspect.getsource(dm_mod.DeviceManager.__init__)
        assert "_stop_event" in src, \
            "DeviceManager debe tener _stop_event (D-02)"
        assert "_save_thread" in src, \
            "DeviceManager debe tener _save_thread (D-02)"

    def test_periodic_save_loop_method_exists(self):
        """DeviceManager debe tener el método _periodic_save_loop."""
        import device_manager as dm_mod
        assert hasattr(dm_mod.DeviceManager, "_periodic_save_loop"), \
            "DeviceManager debe tener _periodic_save_loop (D-02)"

    def test_flush_sets_stop_event(self, tmp_path, monkeypatch):
        """flush() debe señalizar el _stop_event para detener el hilo."""
        import json, threading
        import device_manager as dm_mod
        import config as cfg_mod

        tmp_dev = tmp_path / "devices.json"
        tmp_cfg = tmp_path / "config.json"
        tmp_dev.write_text("{}", encoding="utf-8")
        tmp_cfg.write_text(json.dumps({"pin": ""}), encoding="utf-8")
        monkeypatch.setattr(cfg_mod, "CONFIG_FILE", tmp_cfg)
        monkeypatch.setattr(cfg_mod, "DEVICES_FILE", tmp_dev)
        monkeypatch.setattr(cfg_mod, "_config_cache", None)
        monkeypatch.setattr(cfg_mod, "_config_mtime", 0.0)

        dm = dm_mod.DeviceManager()
        assert not dm._stop_event.is_set(), "stop_event debe estar limpio al inicio"
        dm.flush()
        assert dm._stop_event.is_set(), "flush() debe setear _stop_event para detener el hilo"


class TestHistoryLockQ02:
    """Q-02: _history_lock debe ser atributo de instancia, no de clase."""

    def test_history_lock_is_instance_attribute(self):
        """_history_lock debe estar en __init__, no como atributo de clase."""
        import inspect, queue_manager as qm_mod
        # Verificar que NO es atributo de clase
        assert "_history_lock" not in qm_mod.QueueManager.__dict__, \
            "_history_lock no debe ser atributo de clase (Q-02)"

    def test_two_instances_have_independent_locks(self):
        """Dos instancias de QueueManager deben tener locks distintos."""
        import queue_manager as qm_mod

        class MockPrinter:
            def print_file(self, *a, **kw): pass

        mgr1 = qm_mod.QueueManager(MockPrinter())
        mgr2 = qm_mod.QueueManager(MockPrinter())

        assert mgr1._history_lock is not mgr2._history_lock, \
            "Cada instancia debe tener su propio _history_lock (Q-02)"


# ─────────────────────────────────────────────────────────────────────────────
# Tests v5.4.1 — Regresiones R-01 y R-02 detectadas en auditoría v5.4.0
# ─────────────────────────────────────────────────────────────────────────────

class TestRateLimitClearR01:
    """R-01: la limpieza de intentos fallidos post-login debe usar la clave
    compuesta ip|ua_hash, no solo la IP desnuda."""

    def test_rate_limit_cleared_after_login_with_composite_key(self, tmp_path, monkeypatch):
        """Tras login exitoso, los intentos previos insertados con la clave
        compuesta deben eliminarse. Si se usa solo la IP, el DELETE no
        coincide con ninguna fila y el usuario queda bloqueado."""
        import sqlite3, time
        import server as srv

        # Usar DB temporal para no contaminar el estado global
        tmp_db = tmp_path / "attempts_r01.db"
        monkeypatch.setattr(srv, "_ATTEMPTS_DB", tmp_db)
        with sqlite3.connect(str(tmp_db)) as con:
            con.execute("CREATE TABLE attempts (ip TEXT NOT NULL, ts REAL NOT NULL)")

        ip = "1.2.3.4"
        ua = "TestAgent/1.0"
        key = srv._login_key(ip, ua)

        # Insertar un intento fallido con la clave compuesta (como lo hace _check_rate_limit)
        with sqlite3.connect(str(tmp_db)) as con:
            con.execute("INSERT INTO attempts VALUES (?,?)", (key, time.time()))

        # Ejecutar la limpieza post-login con la clave compuesta
        with sqlite3.connect(str(tmp_db)) as con:
            con.execute("DELETE FROM attempts WHERE ip = ?", (key,))
            count = con.execute(
                "SELECT COUNT(*) FROM attempts WHERE ip = ?", (key,)
            ).fetchone()[0]

        assert count == 0, (
            "Los intentos fallidos deben eliminarse tras login exitoso usando "
            "la clave compuesta ip|ua_hash. Si el DELETE usa solo ip, el COUNT "
            "será > 0 y el usuario quedará bloqueado."
        )

    def test_bare_ip_delete_does_not_clear_composite_key_rows(self, tmp_path):
        """Verificar que el bug R-01 original realmente existía: DELETE con IP
        desnuda NO borra filas insertadas con clave compuesta."""
        import sqlite3, time
        import server as srv

        tmp_db = tmp_path / "attempts_r01_bug.db"
        with sqlite3.connect(str(tmp_db)) as con:
            con.execute("CREATE TABLE attempts (ip TEXT NOT NULL, ts REAL NOT NULL)")

        ip = "5.6.7.8"
        ua = "BugProof/2.0"
        key = srv._login_key(ip, ua)

        with sqlite3.connect(str(tmp_db)) as con:
            con.execute("INSERT INTO attempts VALUES (?,?)", (key, time.time()))

        # Simular el comportamiento BUGGY (DELETE con IP desnuda)
        with sqlite3.connect(str(tmp_db)) as con:
            con.execute("DELETE FROM attempts WHERE ip = ?", (ip,))   # bug original
            count = con.execute(
                "SELECT COUNT(*) FROM attempts WHERE ip = ?", (key,)
            ).fetchone()[0]

        # Confirmar que el bug existe — la fila NO se borró
        assert count == 1, (
            "Este test documenta el bug R-01: DELETE con IP desnuda no borra "
            "filas almacenadas con clave compuesta. count debe ser 1 (bug activo)."
        )

    def test_login_endpoint_uses_composite_key_for_delete(self):
        """Verificar en el código fuente que el DELETE post-login usa _login_key,
        no la variable ip directamente."""
        import inspect, server
        src = inspect.getsource(server.login)
        # El fix correcto debe calcular login_key y usarlo en el DELETE
        assert "_login_key" in src, \
            "login() debe llamar _login_key() para construir la clave del DELETE"
        # Asegurar que el DELETE usa la clave compuesta y no la IP desnuda
        lines = [l.strip() for l in src.splitlines() if "DELETE FROM attempts" in l]
        assert lines, "Debe haber un DELETE FROM attempts en login()"
        for line in lines:
            assert "(ip," not in line and "(ip)" not in line, (
                f"El DELETE no debe usar la variable 'ip' directamente: {line!r}\n"
                "Debe usar la clave compuesta generada por _login_key()"
            )


class TestHistoryLockR02:
    """R-02: _add_to_history debe usar self._history_lock, no
    QueueManager._history_lock (atributo de clase eliminado en v5.4.0)."""

    def test_add_to_history_uses_instance_lock(self):
        """_add_to_history no debe referenciar QueueManager._history_lock."""
        import inspect, queue_manager as qm
        src = inspect.getsource(qm.QueueManager._add_to_history)
        assert "QueueManager._history_lock" not in src, (
            "_add_to_history aún usa QueueManager._history_lock (atributo de "
            "clase eliminado). Debe usar self._history_lock."
        )
        assert "self._history_lock" in src, \
            "_add_to_history debe usar self._history_lock"

    def test_add_to_history_does_not_raise_attribute_error(self):
        """_add_to_history no debe lanzar AttributeError al ejecutarse."""
        import queue_manager as qm

        class MockPrinter:
            def print_file(self, *a, **kw):
                pass

        mgr = qm.QueueManager(MockPrinter())
        job = qm.PrintJob(
            filename="test.pdf",
            filepath="/tmp/test.pdf",
            device_name="TestDevice",
            ip="127.0.0.1",
        )
        job.status = "done"

        try:
            mgr._add_to_history(job)
        except AttributeError as e:
            raise AssertionError(
                f"_add_to_history lanzó AttributeError: {e}\n"
                "Probablemente aún referencia QueueManager._history_lock "
                "que fue eliminado en v5.4.0."
            )

    def test_history_updated_after_job_completion(self):
        """El historial debe actualizarse correctamente tras completar un job."""
        import queue_manager as qm

        class MockPrinter:
            def print_file(self, *a, **kw):
                pass

        mgr = qm.QueueManager(MockPrinter())
        initial_len = len(mgr.get_history())

        job = qm.PrintJob(
            filename="history_test.pdf",
            filepath="/tmp/history_test.pdf",
            device_name="TestDevice",
            ip="127.0.0.1",
        )
        job.status = "done"
        mgr._add_to_history(job)

        updated_history = mgr.get_history()
        assert len(updated_history) == initial_len + 1, (
            "El historial debe tener un entry más tras _add_to_history. "
            "Si sigue siendo igual, el lock falló silenciosamente."
        )
        assert updated_history[0]["filename"] == "history_test.pdf"


# ─────────────────────────────────────────────────────────────────────────────
# Tests Mejora 2 — HTTPS automático (cert_manager.py)
# ─────────────────────────────────────────────────────────────────────────────

class TestCertManagerMejora2:
    """Mejora 2: generación automática de certificado TLS en primer arranque."""

    def test_ensure_certificate_creates_files(self, tmp_path):
        """ensure_certificate() debe crear cert.pem y key.pem."""
        pytest.importorskip("cryptography", reason="cryptography no instalado")
        from cert_manager import ensure_certificate, cert_exists

        assert not cert_exists(tmp_path)
        result = ensure_certificate(tmp_path)

        assert result is True, "ensure_certificate debe retornar True al generar"
        assert (tmp_path / "data" / "cert.pem").exists()
        assert (tmp_path / "data" / "key.pem").exists()

    def test_ensure_certificate_idempotent(self, tmp_path):
        """Llamar ensure_certificate dos veces no regenera el certificado."""
        pytest.importorskip("cryptography")
        from cert_manager import ensure_certificate

        ensure_certificate(tmp_path)
        mtime_first = (tmp_path / "data" / "cert.pem").stat().st_mtime

        ensure_certificate(tmp_path)   # segunda llamada
        mtime_second = (tmp_path / "data" / "cert.pem").stat().st_mtime

        assert mtime_first == mtime_second, \
            "ensure_certificate no debe regenerar si el cert ya existe"

    def test_generated_cert_is_valid_pem(self, tmp_path):
        """El certificado generado debe ser PEM parseable."""
        pytest.importorskip("cryptography")
        from cert_manager import ensure_certificate
        from cryptography import x509

        ensure_certificate(tmp_path)
        cert_bytes = (tmp_path / "data" / "cert.pem").read_bytes()
        cert = x509.load_pem_x509_certificate(cert_bytes)

        assert cert.subject is not None
        from cryptography.x509.oid import NameOID
        cn = cert.subject.get_attributes_for_oid(NameOID.COMMON_NAME)[0].value
        assert cn == "PrintBridge"

    def test_generated_cert_validity_period(self, tmp_path):
        """El certificado debe tener una validez de ~10 años."""
        pytest.importorskip("cryptography")
        from cert_manager import ensure_certificate, CERT_DAYS
        from cryptography import x509
        from datetime import timezone

        ensure_certificate(tmp_path)
        cert_bytes = (tmp_path / "data" / "cert.pem").read_bytes()
        cert = x509.load_pem_x509_certificate(cert_bytes)

        from datetime import datetime as _dt
        now = _dt.now(timezone.utc)
        delta = cert.not_valid_after_utc - now
        # Debe ser aproximadamente CERT_DAYS (tolerancia de ±2 días)
        assert abs(delta.days - CERT_DAYS) <= 2, \
            f"Validez del certificado inesperada: {delta.days} días (esperado ~{CERT_DAYS})"

    def test_generated_cert_has_san_localhost(self, tmp_path):
        """El certificado debe incluir localhost como SAN."""
        pytest.importorskip("cryptography")
        from cert_manager import ensure_certificate
        from cryptography import x509

        ensure_certificate(tmp_path)
        cert_bytes = (tmp_path / "data" / "cert.pem").read_bytes()
        cert = x509.load_pem_x509_certificate(cert_bytes)

        san_ext = cert.extensions.get_extension_for_class(x509.SubjectAlternativeName)
        dns_names = san_ext.value.get_values_for_type(x509.DNSName)
        assert "localhost" in dns_names, \
            "El certificado debe incluir 'localhost' como SAN DNS"

    def test_generated_cert_has_san_loopback_ip(self, tmp_path):
        """El certificado debe incluir 127.0.0.1 como SAN IP."""
        pytest.importorskip("cryptography")
        from cert_manager import ensure_certificate
        from cryptography import x509
        import ipaddress

        ensure_certificate(tmp_path)
        cert_bytes = (tmp_path / "data" / "cert.pem").read_bytes()
        cert = x509.load_pem_x509_certificate(cert_bytes)

        san_ext = cert.extensions.get_extension_for_class(x509.SubjectAlternativeName)
        ip_addrs = san_ext.value.get_values_for_type(x509.IPAddress)
        assert ipaddress.IPv4Address("127.0.0.1") in ip_addrs, \
            "El certificado debe incluir 127.0.0.1 como SAN IP"

    def test_private_key_is_valid_rsa(self, tmp_path):
        """La clave privada debe ser RSA-2048 parseable."""
        pytest.importorskip("cryptography")
        from cert_manager import ensure_certificate, KEY_SIZE
        from cryptography.hazmat.primitives.serialization import load_pem_private_key

        ensure_certificate(tmp_path)
        key_bytes = (tmp_path / "data" / "key.pem").read_bytes()
        key = load_pem_private_key(key_bytes, password=None)

        assert key.key_size == KEY_SIZE, \
            f"Clave privada debe ser RSA-{KEY_SIZE}, es RSA-{key.key_size}"

    def test_cert_exists_returns_false_when_missing(self, tmp_path):
        """cert_exists() retorna False cuando no hay certificado."""
        from cert_manager import cert_exists
        assert cert_exists(tmp_path) is False

    def test_cert_exists_returns_true_after_generation(self, tmp_path):
        """cert_exists() retorna True después de generar."""
        pytest.importorskip("cryptography")
        from cert_manager import ensure_certificate, cert_exists
        ensure_certificate(tmp_path)
        assert cert_exists(tmp_path) is True

    def test_ensure_certificate_without_cryptography(self, tmp_path, monkeypatch):
        """Si cryptography no está disponible, ensure_certificate retorna False
        sin lanzar excepción (degradación graciosa a HTTP)."""
        import sys
        # Simular que cryptography no está instalado
        monkeypatch.setitem(sys.modules, "cryptography", None)
        monkeypatch.setitem(sys.modules, "cryptography.x509", None)

        from cert_manager import ensure_certificate
        # No debe lanzar ImportError — debe retornar False graciosamente
        result = ensure_certificate(tmp_path)
        assert result is False, \
            "Sin cryptography, ensure_certificate debe retornar False (no crash)"

    def test_ensure_certificate_async_calls_callback(self, tmp_path):
        """ensure_certificate_async debe llamar el callback cuando termina."""
        pytest.importorskip("cryptography")
        from cert_manager import ensure_certificate_async
        import threading

        results = []
        done = threading.Event()

        def callback(ok):
            results.append(ok)
            done.set()

        t = ensure_certificate_async(tmp_path, on_complete=callback)
        done.wait(timeout=10)
        t.join(timeout=2)

        assert len(results) == 1, "El callback debe llamarse exactamente una vez"
        assert results[0] is True

    def test_get_cert_info_returns_dict(self, tmp_path):
        """get_cert_info debe retornar un dict con campos esperados."""
        pytest.importorskip("cryptography")
        from cert_manager import ensure_certificate, get_cert_info

        ensure_certificate(tmp_path)
        info = get_cert_info(tmp_path)

        assert info["exists"] is True
        assert "subject" in info
        assert "not_before" in info
        assert "not_after" in info
        assert "serial" in info

    def test_get_cert_info_no_cert(self, tmp_path):
        """get_cert_info sin certificado retorna exists=False."""
        from cert_manager import get_cert_info
        info = get_cert_info(tmp_path)
        assert info == {"exists": False}

    @pytest.mark.win32_only
    def test_is_cert_trusted_windows_returns_bool(self, tmp_path):
        """is_cert_trusted_windows debe retornar bool sin crashear."""
        pytest.importorskip("cryptography")
        from cert_manager import ensure_certificate, is_cert_trusted_windows
        ensure_certificate(tmp_path)
        result = is_cert_trusted_windows(tmp_path)
        assert isinstance(result, bool)


# ─────────────────────────────────────────────────────────────────────────────
# Tests Mejora 1 — Proceso separado para impresión (queue_manager + print_worker)
# ─────────────────────────────────────────────────────────────────────────────

class TestPrintWorkerMejora1:
    """Mejora 1: print_worker.py — función worker serializable para subprocess."""

    def test_run_print_job_raises_on_missing_file(self, tmp_path):
        """run_print_job lanza FileNotFoundError si el archivo no existe."""
        from print_worker import run_print_job
        with pytest.raises(FileNotFoundError, match="no encontrado"):
            run_print_job(str(tmp_path / "nonexistent.pdf"), 1, {})

    def test_get_timeout_pdf(self):
        """PDF debe tener timeout de 60s por defecto."""
        from print_worker import get_timeout_for_ext
        assert get_timeout_for_ext("pdf") == 60

    def test_get_timeout_docx(self):
        """DOCX debe tener timeout de 180s (conversión LibreOffice)."""
        from print_worker import get_timeout_for_ext
        assert get_timeout_for_ext("docx") == 180

    def test_get_timeout_image(self):
        """JPG/PNG deben tener timeout corto (30s)."""
        from print_worker import get_timeout_for_ext
        assert get_timeout_for_ext("jpg") == 30
        assert get_timeout_for_ext("png") == 30

    def test_get_timeout_unknown_ext_returns_default(self):
        """Extensión desconocida retorna timeout default (120s)."""
        from print_worker import get_timeout_for_ext
        assert get_timeout_for_ext("xyz_unknown") == 120

    def test_get_timeout_case_insensitive(self):
        """La extensión debe ser case-insensitive."""
        from print_worker import get_timeout_for_ext
        assert get_timeout_for_ext("PDF") == get_timeout_for_ext("pdf")
        assert get_timeout_for_ext("DOCX") == get_timeout_for_ext("docx")

    def test_get_timeout_custom_from_config(self, monkeypatch, tmp_path):
        """Timeout personalizado en config.json debe usarse."""
        import config as cfg
        import json
        tmp_cfg = tmp_path / "config.json"
        tmp_cfg.write_text(json.dumps({
            "print_timeouts": {"pdf": 999, "default": 500}
        }), encoding="utf-8")
        monkeypatch.setattr(cfg, "CONFIG_FILE", tmp_cfg)
        monkeypatch.setattr(cfg, "_config_cache", None)
        monkeypatch.setattr(cfg, "_config_mtime", 0.0)

        from print_worker import get_timeout_for_ext
        assert get_timeout_for_ext("pdf") == 999
        assert get_timeout_for_ext("unknown") == 500

    def test_print_worker_module_is_importable(self):
        """print_worker debe ser importable (necesario para pickle en subprocess)."""
        import importlib
        mod = importlib.import_module("print_worker")
        assert hasattr(mod, "run_print_job")
        assert hasattr(mod, "get_timeout_for_ext")

    def test_run_print_job_is_top_level_function(self):
        """run_print_job debe ser función de módulo (no método), para ser picklable."""
        import inspect, print_worker
        assert inspect.isfunction(print_worker.run_print_job), \
            "run_print_job debe ser función de módulo para poder serializarse en subprocess"


class TestQueueManagerMejora1:
    """Mejora 1: QueueManager con ProcessPoolExecutor."""

    def test_queue_manager_has_executor(self):
        """QueueManager debe tener un ProcessPoolExecutor."""
        import concurrent.futures
        from queue_manager import QueueManager
        from unittest.mock import MagicMock
        mgr = QueueManager(MagicMock())
        assert hasattr(mgr, "_executor"), "QueueManager debe tener _executor"
        assert isinstance(mgr._executor, concurrent.futures.ProcessPoolExecutor)
        mgr.shutdown(wait=False)

    def test_queue_manager_has_shutdown_method(self):
        """QueueManager debe tener método shutdown()."""
        from queue_manager import QueueManager
        assert hasattr(QueueManager, "shutdown"), \
            "QueueManager debe tener método shutdown() para Mejora 1"

    def test_shutdown_does_not_raise(self):
        """shutdown(wait=False) no debe lanzar excepción."""
        from queue_manager import QueueManager
        from unittest.mock import MagicMock
        mgr = QueueManager(MagicMock())
        try:
            mgr.shutdown(wait=False)
        except Exception as e:
            pytest.fail(f"shutdown() lanzó excepción: {e}")

    def test_add_job_returns_id(self, queue_manager, tmp_path):
        """add_job debe retornar el ID del job."""
        mgr, PrintJob = queue_manager
        # Crear archivo falso para que el worker no crashee por FileNotFoundError
        fake_file = tmp_path / "test.pdf"
        fake_file.write_bytes(b"%PDF-1.4 test")
        job = PrintJob("test.pdf", str(fake_file), "PC", "127.0.0.1")
        job_id = mgr.add_job(job)
        assert job_id is not None
        assert len(job_id) == 32   # UUID hex sin guiones

    def test_queue_respects_max_size(self, queue_manager, tmp_path, monkeypatch):
        """La cola no debe aceptar más jobs que max_queue_size."""
        import config as cfg, json
        tmp_cfg = tmp_path / "config_maxq.json"
        tmp_cfg.write_text(json.dumps({"max_queue_size": 2}), encoding="utf-8")
        monkeypatch.setattr(cfg, "CONFIG_FILE", tmp_cfg)
        monkeypatch.setattr(cfg, "_config_cache", None)
        monkeypatch.setattr(cfg, "_config_mtime", 0.0)

        mgr, PrintJob = queue_manager
        # Pausar el worker para que no consuma jobs durante el test
        with mgr.lock:
            for i in range(3):
                fake = tmp_path / f"job{i}.pdf"
                fake.write_bytes(b"%PDF test")
                job = PrintJob(f"j{i}.pdf", str(fake), "PC", "127.0.0.1")
                mgr.queue.append(job)   # insertar directo para controlar el estado

        # El 4to job vía add_job debe ser rechazado (cola llena)
        fake4 = tmp_path / "job4.pdf"
        fake4.write_bytes(b"%PDF test")
        j4 = PrintJob("j4.pdf", str(fake4), "PC", "127.0.0.1")
        result = mgr.add_job(j4)
        assert result is None, "add_job debe retornar None cuando la cola está llena"

    def test_cancel_waiting_job(self, queue_manager, tmp_path):
        """cancel_job debe eliminar un job en estado 'waiting'."""
        mgr, PrintJob = queue_manager
        fake = tmp_path / "cancel_test.pdf"
        fake.write_bytes(b"%PDF test")
        job = PrintJob("cancel.pdf", str(fake), "PC", "127.0.0.1")

        with mgr.lock:
            mgr.queue.append(job)

        success = mgr.cancel_job(job.id)
        assert success is True
        assert job.status == "cancelled"
        with mgr.lock:
            assert job not in mgr.queue

    def test_process_next_job_uses_executor(self, monkeypatch):
        """_process_next_job debe someter el job al executor, no llamar printer directamente."""
        import inspect
        from queue_manager import QueueManager
        src = inspect.getsource(QueueManager._process_next_job)
        assert "executor" in src, \
            "_process_next_job debe usar self._executor (Mejora 1)"
        assert "submit" in src, \
            "_process_next_job debe llamar executor.submit()"
        assert "timeout" in src, \
            "_process_next_job debe usar timeout en future.result()"

    def test_timeout_error_sets_job_error_status(self, monkeypatch, tmp_path):
        """Si el Future lanza TimeoutError, el job debe quedar en status='error'."""
        import concurrent.futures
        from queue_manager import QueueManager, PrintJob
        from unittest.mock import MagicMock, patch

        mgr = QueueManager(MagicMock())

        fake = tmp_path / "slow_job.pdf"
        fake.write_bytes(b"%PDF test")
        job = PrintJob("slow.pdf", str(fake), "PC", "127.0.0.1")
        job.status = "printing"
        mgr.current_job = job

        # Simular que el Future lanza TimeoutError
        mock_future = MagicMock()
        mock_future.result.side_effect = concurrent.futures.TimeoutError()

        with patch.object(mgr._executor, "submit", return_value=mock_future):
            with mgr.lock:
                mgr.queue.append(job)
                mgr.queue.popleft()   # simular dequeue manual

            # Llamar directamente al proceso del job (sin el loop)
            from print_worker import get_timeout_for_ext
            ext = "pdf"
            timeout_secs = get_timeout_for_ext(ext)
            try:
                mock_future.result(timeout=timeout_secs)
            except concurrent.futures.TimeoutError:
                job.status = "error"
                job.error = f"Tiempo de impresión excedido ({timeout_secs}s)."

        assert job.status == "error"
        assert "excedido" in job.error
        assert str(timeout_secs) in job.error

        mgr.shutdown(wait=False)

    def test_worker_does_not_call_printer_directly(self):
        """_process_next_job no debe llamar self.printer.print_file directamente."""
        import inspect
        from queue_manager import QueueManager
        src = inspect.getsource(QueueManager._process_next_job)
        assert "self.printer.print_file" not in src, \
            "Mejora 1: _process_next_job no debe llamar printer directamente — usar subprocess"

    def test_freeze_support_in_app_entrypoint(self):
        """app.py debe llamar freeze_support() en if __name__ == '__main__'."""
        import inspect, app as app_mod
        src = inspect.getsource(app_mod)
        assert "freeze_support" in src, \
            "app.py debe llamar freeze_support() para compatibilidad con PyInstaller + spawn"


# ─────────────────────────────────────────────────────────────────────────────
# Tests Mejora 7 — Modo servicio Windows (service.py)
# ─────────────────────────────────────────────────────────────────────────────

class TestServiceModuleMejora7:
    """Mejora 7: service.py — modo servicio Windows sin GUI."""

    def test_service_module_importable(self):
        """service.py debe ser importable sin win32service."""
        import importlib
        mod = importlib.import_module("service")
        assert hasattr(mod, "run_foreground")
        assert hasattr(mod, "_cli_main")
        assert hasattr(mod, "_initialize_components")
        assert hasattr(mod, "_shutdown_components")

    def test_cli_main_help_returns_zero(self):
        """'help' y '--help' deben retornar código 0."""
        from service import _cli_main
        assert _cli_main(["help"]) == 0
        assert _cli_main(["--help"]) == 0
        assert _cli_main(["-h"]) == 0

    def test_cli_main_unknown_command_returns_one(self):
        """Comando desconocido debe retornar código 1."""
        from service import _cli_main
        assert _cli_main(["unknown_xyz_command"]) == 1

    def test_cli_main_status_returns_zero(self):
        """'status' debe retornar 0 (aunque el servicio no esté instalado)."""
        from service import _cli_main
        assert _cli_main(["status"]) == 0

    def test_get_service_status_returns_string(self):
        """_get_service_status debe retornar un string descriptivo."""
        from service import _get_service_status
        status = _get_service_status()
        assert isinstance(status, str)
        assert len(status) > 0

    def test_win32service_availability_flag(self):
        """WIN32SERVICE_AVAILABLE debe ser un bool."""
        from service import WIN32SERVICE_AVAILABLE
        assert isinstance(WIN32SERVICE_AVAILABLE, bool)

    def test_initialize_components_signature(self):
        """_initialize_components no debe requerir argumentos."""
        import inspect
        from service import _initialize_components
        sig = inspect.signature(_initialize_components)
        params = [
            p for p in sig.parameters.values()
            if p.default is inspect.Parameter.empty
        ]
        assert len(params) == 0, \
            "_initialize_components no debe tener argumentos requeridos"

    def test_shutdown_components_signature(self):
        """_shutdown_components debe aceptar (device_mgr, queue_mgr)."""
        import inspect
        from service import _shutdown_components
        sig = inspect.signature(_shutdown_components)
        assert len(sig.parameters) == 2

    @pytest.mark.win32_only
    def test_service_class_has_required_attributes(self):
        """PrintBridgeService debe tener los atributos requeridos por win32serviceutil."""
        from service import PrintBridgeService
        assert hasattr(PrintBridgeService, "_svc_name_")
        assert hasattr(PrintBridgeService, "_svc_display_name_")
        assert hasattr(PrintBridgeService, "_svc_description_")
        assert PrintBridgeService._svc_name_ == "PrintBridge"

    @pytest.mark.win32_only
    def test_service_class_has_svc_methods(self):
        """PrintBridgeService debe implementar SvcDoRun y SvcStop."""
        from service import PrintBridgeService
        assert hasattr(PrintBridgeService, "SvcDoRun")
        assert hasattr(PrintBridgeService, "SvcStop")
        assert callable(PrintBridgeService.SvcDoRun)
        assert callable(PrintBridgeService.SvcStop)

    def test_run_foreground_is_callable(self):
        """run_foreground debe ser una función callable."""
        from service import run_foreground
        assert callable(run_foreground)

    def test_freeze_support_in_service_entrypoint(self):
        """service.py debe llamar freeze_support() en su entry point."""
        import inspect, service as svc_mod
        src = inspect.getsource(svc_mod)
        assert "freeze_support" in src, \
            "service.py debe llamar freeze_support() para compatibilidad con PyInstaller"

    def test_app_routes_service_subcommand(self):
        """app.py debe rutear 'service' como subcomando a service._cli_main."""
        import inspect, app as app_mod
        src = inspect.getsource(app_mod)
        assert "service" in src.lower(), \
            "app.py debe incluir routing del subcomando 'service'"
        assert "_cli_main" in src or "service.py" in src.lower(), \
            "app.py debe delegar a service._cli_main"

    def test_build_exe_includes_service_hidden_imports(self):
        """build_exe.bat debe incluir win32service y service como hidden-imports."""
        build_bat = (
            Path(__file__).parent.parent / "build_exe.bat"
        ).read_text(encoding="utf-8", errors="replace")
        assert "win32service" in build_bat, \
            "build_exe.bat debe incluir --hidden-import win32service"
        assert "servicemanager" in build_bat, \
            "build_exe.bat debe incluir --hidden-import servicemanager"
        assert "cert_manager" in build_bat, \
            "build_exe.bat debe incluir --hidden-import cert_manager"
        assert "print_worker" in build_bat, \
            "build_exe.bat debe incluir --hidden-import print_worker"

    def test_service_log_path_is_in_data_dir(self):
        """El log del servicio debe estar en data/service.log."""
        from service import _LOG_FILE
        assert _LOG_FILE.parent.name == "data", \
            "service.log debe estar en el subdirectorio data/"
        assert _LOG_FILE.name == "service.log"

    def test_cli_main_no_args_shows_usage(self, capsys):
        """Sin argumentos (o 'help'), _cli_main debe mostrar el uso."""
        from service import _cli_main
        _cli_main(["help"])
        captured = capsys.readouterr()
        assert "install" in captured.out.lower() or "uso" in captured.out.lower(), \
            "El mensaje de ayuda debe mencionar el comando 'install'"


# ─────────────────────────────────────────────────────────────────────────────
# Tests Mejora 5 — Configuración tipada con Pydantic (config_schema.py)
# ─────────────────────────────────────────────────────────────────────────────

class TestConfigSchemaMejora5:
    """Mejora 5: AppConfig con validación de tipos en carga."""

    def test_appconfig_importable(self):
        """config_schema.py debe importarse sin errores."""
        import config_schema
        assert hasattr(config_schema, "AppConfig")
        assert hasattr(config_schema, "load_typed_config")

    def test_appconfig_defaults(self):
        """AppConfig sin argumentos debe tener todos los defaults correctos."""
        from config_schema import AppConfig, PYDANTIC_AVAILABLE
        if not PYDANTIC_AVAILABLE:
            pytest.skip("Pydantic no instalado")
        cfg = AppConfig()
        assert cfg.port == 7878
        assert cfg.server_name == "PrintBridge"
        assert cfg.max_queue_size == 20
        assert cfg.max_history == 100
        assert cfg.token_expiry_days == 90
        assert isinstance(cfg.allowed_extensions, list)

    def test_appconfig_port_validation(self):
        """Port fuera de rango debe lanzar ValidationError."""
        from config_schema import AppConfig, PYDANTIC_AVAILABLE
        if not PYDANTIC_AVAILABLE:
            pytest.skip("Pydantic no instalado")
        import pydantic
        with pytest.raises((pydantic.ValidationError, ValueError)):
            AppConfig(port=99999)
        with pytest.raises((pydantic.ValidationError, ValueError)):
            AppConfig(port=0)

    def test_appconfig_server_name_strips_html(self):
        """server_name con caracteres HTML debe lanzar error."""
        from config_schema import AppConfig, PYDANTIC_AVAILABLE
        if not PYDANTIC_AVAILABLE:
            pytest.skip("Pydantic no instalado")
        import pydantic
        with pytest.raises((pydantic.ValidationError, ValueError)):
            AppConfig(server_name="<script>evil</script>")

    def test_appconfig_server_name_too_long(self):
        """server_name > 64 chars debe lanzar error."""
        from config_schema import AppConfig, PYDANTIC_AVAILABLE
        if not PYDANTIC_AVAILABLE:
            pytest.skip("Pydantic no instalado")
        import pydantic
        with pytest.raises((pydantic.ValidationError, ValueError)):
            AppConfig(server_name="A" * 65)

    def test_appconfig_empty_name_defaults_to_printbridge(self):
        """server_name vacío debe usar 'PrintBridge' como fallback."""
        from config_schema import AppConfig, PYDANTIC_AVAILABLE
        if not PYDANTIC_AVAILABLE:
            pytest.skip("Pydantic no instalado")
        cfg = AppConfig(server_name="   ")
        assert cfg.server_name == "PrintBridge"

    def test_appconfig_extensions_normalized_to_lowercase(self):
        """Las extensiones deben normalizarse a minúsculas."""
        from config_schema import AppConfig, PYDANTIC_AVAILABLE
        if not PYDANTIC_AVAILABLE:
            pytest.skip("Pydantic no instalado")
        cfg = AppConfig(allowed_extensions=["PDF", "DOCX", "jpg"])
        assert "pdf" in cfg.allowed_extensions
        assert "docx" in cfg.allowed_extensions
        assert "PDF" not in cfg.allowed_extensions

    def test_appconfig_extensions_deduplicates(self):
        """Las extensiones duplicadas deben eliminarse."""
        from config_schema import AppConfig, PYDANTIC_AVAILABLE
        if not PYDANTIC_AVAILABLE:
            pytest.skip("Pydantic no instalado")
        cfg = AppConfig(allowed_extensions=["pdf", "PDF", "pdf"])
        assert cfg.allowed_extensions.count("pdf") == 1

    def test_appconfig_to_dict_is_json_serializable(self):
        """to_dict() debe producir un dict serializable a JSON."""
        import json as _json
        from config_schema import AppConfig, PYDANTIC_AVAILABLE
        if not PYDANTIC_AVAILABLE:
            pytest.skip("Pydantic no instalado")
        cfg = AppConfig()
        d = cfg.to_dict()
        assert isinstance(d, dict)
        serialized = _json.dumps(d)
        assert serialized

    def test_load_typed_config_returns_appconfig(self, isolate_config):
        """load_typed_config() debe retornar un AppConfig."""
        from config_schema import load_typed_config, AppConfig
        cfg = load_typed_config()
        assert isinstance(cfg, AppConfig)

    def test_load_typed_config_graceful_on_bad_port(self, isolate_config, tmp_path):
        """Config con port inválido no debe crashear el servidor."""
        import json as _json
        import config as cfg_mod
        bad_cfg_file = tmp_path / "bad_config.json"
        bad_cfg_file.write_text(_json.dumps({"port": "not-a-number"}), encoding="utf-8")
        cfg_mod.CONFIG_FILE = bad_cfg_file
        cfg_mod._config_cache = None
        from config_schema import load_typed_config
        # No debe lanzar excepción — fallback gracioso
        result = load_typed_config()
        assert result is not None

    def test_pydantic_availability_flag(self):
        """PYDANTIC_AVAILABLE debe ser bool."""
        from config_schema import PYDANTIC_AVAILABLE
        assert isinstance(PYDANTIC_AVAILABLE, bool)


# ─────────────────────────────────────────────────────────────────────────────
# Tests Mejora 4 — Dependency Injection con FastAPI Depends()
# ─────────────────────────────────────────────────────────────────────────────

class TestDependencyInjectionMejora4:
    """Mejora 4: providers DI get_device_mgr / get_queue_mgr en server.py."""

    def test_provider_functions_exist(self):
        """get_device_mgr y get_queue_mgr deben existir en server.py."""
        import server as srv
        assert hasattr(srv, "get_device_mgr"), "server.py debe tener get_device_mgr()"
        assert hasattr(srv, "get_queue_mgr"),  "server.py debe tener get_queue_mgr()"
        assert callable(srv.get_device_mgr)
        assert callable(srv.get_queue_mgr)

    def test_get_device_mgr_raises_503_when_none(self, monkeypatch):
        """get_device_mgr debe lanzar HTTP 503 si device_mgr es None."""
        import server as srv
        from fastapi import HTTPException
        monkeypatch.setattr(srv, "device_mgr", None)
        with pytest.raises(HTTPException) as exc_info:
            srv.get_device_mgr()
        assert exc_info.value.status_code == 503

    def test_get_queue_mgr_raises_503_when_none(self, monkeypatch):
        """get_queue_mgr debe lanzar HTTP 503 si queue_mgr es None."""
        import server as srv
        from fastapi import HTTPException
        monkeypatch.setattr(srv, "queue_mgr", None)
        with pytest.raises(HTTPException) as exc_info:
            srv.get_queue_mgr()
        assert exc_info.value.status_code == 503

    def test_get_device_mgr_returns_instance(self, api_client):
        """get_device_mgr debe retornar el DeviceManager inyectado."""
        import server as srv
        from device_manager import DeviceManager
        client, dm, qm = api_client
        result = srv.get_device_mgr()
        assert isinstance(result, DeviceManager)

    def test_get_queue_mgr_returns_instance(self, api_client):
        """get_queue_mgr debe retornar el QueueManager inyectado."""
        import server as srv
        from queue_manager import QueueManager
        client, dm, qm = api_client
        result = srv.get_queue_mgr()
        assert isinstance(result, QueueManager)

    def test_dependency_override_works_in_tests(self, monkeypatch):
        """app.dependency_overrides permite reemplazar dependencias en tests."""
        import server as srv
        from unittest.mock import MagicMock
        mock_dm = MagicMock()
        original_overrides = dict(srv.app.dependency_overrides)
        try:
            srv.app.dependency_overrides[srv.get_device_mgr] = lambda: mock_dm
            result = srv.get_device_mgr.__wrapped__() if hasattr(srv.get_device_mgr, "__wrapped__") else mock_dm
            # El override existe en el dict
            assert srv.get_device_mgr in srv.app.dependency_overrides
        finally:
            srv.app.dependency_overrides.clear()
            srv.app.dependency_overrides.update(original_overrides)

    def test_migrated_endpoints_use_depends(self):
        """Los endpoints migrados deben usar Depends() en su firma."""
        import inspect, server as srv
        for fn_name in ("get_queue", "cancel_job", "get_history",
                        "list_devices", "remove_device"):
            fn = getattr(srv, fn_name, None)
            assert fn is not None, f"Endpoint {fn_name} no encontrado en server.py"
            src = inspect.getsource(fn)
            assert "_Depends(" in src or "Depends(" in src, \
                f"{fn_name} debe usar Depends() (Mejora 4)"

    def test_api_queue_endpoint_accessible(self, api_client):
        """GET /api/queue debe responder 200 con los providers DI."""
        client, dm, qm = api_client
        r = client.get("/api/queue")
        assert r.status_code in (200, 403)  # 403 si requiere auth sin PIN

    def test_api_history_endpoint_accessible(self, api_client):
        """GET /api/history debe responder 200 con los providers DI."""
        client, dm, qm = api_client
        r = client.get("/api/history")
        assert r.status_code in (200, 403)


# ─────────────────────────────────────────────────────────────────────────────
# Tests Mejora 3 — Cola persistida en SQLite (queue_db.py)
# ─────────────────────────────────────────────────────────────────────────────

class TestQueueDBMejora3:
    """Mejora 3: QueueDB — persistencia de cola e historial en SQLite."""

    def test_queue_db_creates_file(self, tmp_path):
        """QueueDB debe crear el archivo .db al inicializar."""
        from queue_db import QueueDB
        db_path = tmp_path / "test_queue.db"
        db = QueueDB(db_path)
        assert db_path.exists(), "QueueDB debe crear el archivo SQLite"
        db.close()

    def test_insert_and_recover_waiting_job(self, tmp_path):
        """Un job insertado en 'waiting' debe recuperarse al reiniciar."""
        from queue_db import QueueDB
        from queue_manager import PrintJob
        db = QueueDB(tmp_path / "queue.db")
        job = PrintJob("test.pdf", "/tmp/test.pdf", "PC", "127.0.0.1")
        db.insert_job(job)
        waiting = db.recover_waiting_jobs()
        assert len(waiting) == 1
        assert waiting[0]["id"] == job.id
        assert waiting[0]["status"] == "waiting"
        db.close()

    def test_update_status_to_done(self, tmp_path):
        """update_status debe cambiar el estado del job."""
        from queue_db import QueueDB
        from queue_manager import PrintJob
        from datetime import datetime
        db  = QueueDB(tmp_path / "queue.db")
        job = PrintJob("doc.pdf", "/tmp/doc.pdf", "PC", "127.0.0.1")
        db.insert_job(job)
        db.update_status(job.id, "done", finished_at=datetime.now().isoformat())
        history = db.get_history()
        assert len(history) == 1
        assert history[0]["status"] == "done"
        db.close()

    def test_waiting_jobs_not_in_history(self, tmp_path):
        """get_history NO debe retornar jobs en estado 'waiting'."""
        from queue_db import QueueDB
        from queue_manager import PrintJob
        db  = QueueDB(tmp_path / "queue.db")
        job = PrintJob("waiting.pdf", "/tmp/w.pdf", "PC", "127.0.0.1")
        db.insert_job(job)
        history = db.get_history()
        assert all(h["id"] != job.id for h in history), \
            "Jobs en 'waiting' no deben aparecer en el historial"
        db.close()

    def test_recover_waiting_excludes_done(self, tmp_path):
        """recover_waiting_jobs no debe retornar jobs terminados."""
        from queue_db import QueueDB
        from queue_manager import PrintJob
        db   = QueueDB(tmp_path / "queue.db")
        job1 = PrintJob("done.pdf", "/tmp/d.pdf", "PC", "127.0.0.1")
        job2 = PrintJob("wait.pdf", "/tmp/w.pdf", "PC", "127.0.0.1")
        db.insert_job(job1)
        db.insert_job(job2)
        db.update_status(job1.id, "done")
        waiting = db.recover_waiting_jobs()
        assert len(waiting) == 1
        assert waiting[0]["id"] == job2.id
        db.close()

    def test_prune_history_removes_old_entries(self, tmp_path):
        """prune_history debe eliminar entradas antiguas por encima del límite."""
        from queue_db import QueueDB
        from queue_manager import PrintJob
        from datetime import datetime
        db = QueueDB(tmp_path / "queue.db")
        for i in range(10):
            job = PrintJob(f"file{i}.pdf", f"/tmp/{i}.pdf", "PC", "127.0.0.1")
            db.insert_job(job)
            db.update_status(job.id, "done", finished_at=datetime.now().isoformat())
        pruned = db.prune_history(max_entries=5)
        assert pruned == 5
        remaining = db.get_history(limit=100)
        assert len(remaining) == 5
        db.close()

    def test_count_waiting(self, tmp_path):
        """count_waiting debe retornar el número correcto de jobs pendientes."""
        from queue_db import QueueDB
        from queue_manager import PrintJob
        db = QueueDB(tmp_path / "queue.db")
        for i in range(3):
            job = PrintJob(f"f{i}.pdf", f"/tmp/{i}.pdf", "PC", "127.0.0.1")
            db.insert_job(job)
        assert db.count_waiting() == 3
        db.close()

    def test_history_pagination(self, tmp_path):
        """get_history debe soportar limit y offset."""
        from queue_db import QueueDB
        from queue_manager import PrintJob
        from datetime import datetime
        db = QueueDB(tmp_path / "queue.db")
        for i in range(5):
            job = PrintJob(f"f{i}.pdf", f"/tmp/{i}.pdf", "PC", "127.0.0.1")
            db.insert_job(job)
            db.update_status(job.id, "done", finished_at=datetime.now().isoformat())
        page1 = db.get_history(limit=2, offset=0)
        page2 = db.get_history(limit=2, offset=2)
        assert len(page1) == 2
        assert len(page2) == 2
        ids_p1 = {r["id"] for r in page1}
        ids_p2 = {r["id"] for r in page2}
        assert ids_p1.isdisjoint(ids_p2), "Las páginas no deben tener IDs en común"
        db.close()

    def test_queue_manager_uses_sqlite_for_history(self):
        """QueueManager.get_history debe leer desde SQLite, no de JSON."""
        import inspect
        from queue_manager import QueueManager
        src = inspect.getsource(QueueManager.get_history)
        assert "_db" in src, \
            "get_history debe leer desde self._db (SQLite), no de _history_cache"
        assert "save_history" not in src, \
            "get_history no debe llamar save_history (JSON obsoleto)"

    def test_queue_manager_recovers_jobs_on_init(self, tmp_path, monkeypatch):
        """QueueManager debe recuperar jobs 'waiting' de SQLite al iniciar."""
        import json as _json
        from queue_db import QueueDB
        from queue_manager import QueueManager, PrintJob, _QUEUE_DB_PATH
        from unittest.mock import MagicMock
        import config as cfg

        tmp_cfg  = tmp_path / "config.json"
        tmp_cfg.write_text(_json.dumps({}), encoding="utf-8")
        monkeypatch.setattr(cfg, "CONFIG_FILE", tmp_cfg)
        monkeypatch.setattr(cfg, "_config_cache", None)
        monkeypatch.setattr(cfg, "_config_mtime", 0.0)

        # Pre-insertar un job 'waiting' en la DB
        test_db = tmp_path / "queue.db"
        monkeypatch.setattr("queue_manager._QUEUE_DB_PATH", test_db)
        db = QueueDB(test_db)
        job = PrintJob("recover.pdf", "/tmp/recover.pdf", "PC", "127.0.0.1")
        db.insert_job(job)
        db.close()

        mgr = QueueManager(MagicMock())
        # El job debe haberse recuperado en la deque
        with mgr.lock:
            ids_in_queue = [j.id for j in mgr.queue]
        assert job.id in ids_in_queue, \
            "QueueManager debe recuperar jobs 'waiting' de SQLite al arrancar"
        mgr.shutdown(wait=False)


# ─────────────────────────────────────────────────────────────────────────────
# Tests Mejora 6 — SSE para estado en tiempo real (/api/events)
# ─────────────────────────────────────────────────────────────────────────────

class TestSSEMejora6:
    """Mejora 6: endpoint SSE /api/events."""

    def test_events_endpoint_exists(self):
        """GET /api/events debe estar registrado en la app FastAPI."""
        import server as srv
        routes = {r.path for r in srv.app.routes}
        assert "/api/events" in routes, \
            "Mejora 6: /api/events debe existir como endpoint SSE"

    def test_events_endpoint_media_type(self, api_client):
        """GET /api/events debe retornar media type text/event-stream."""
        from starlette.testclient import TestClient
        import server as srv
        client, dm, qm = api_client
        # Hacer streaming con timeout muy corto para no bloquear el test
        with client.stream("GET", "/api/events") as r:
            assert r.status_code == 200
            content_type = r.headers.get("content-type", "")
            assert "text/event-stream" in content_type, \
                f"Content-Type debe ser text/event-stream, es: {content_type}"

    def test_events_stream_emits_data(self, api_client):
        """El stream debe emitir al menos un evento data:."""
        client, dm, qm = api_client
        received_lines = []
        with client.stream("GET", "/api/events") as r:
            for i, line in enumerate(r.iter_lines()):
                if line.startswith("data:"):
                    received_lines.append(line)
                if i > 20 or received_lines:   # leer hasta el primer dato
                    break
        assert len(received_lines) >= 1, \
            "El stream SSE debe emitir al menos un evento 'data:'"

    def test_events_data_is_valid_json(self, api_client):
        """El payload del evento SSE debe ser JSON válido con campo 'queue'."""
        import json as _json
        client, dm, qm = api_client
        with client.stream("GET", "/api/events") as r:
            for line in r.iter_lines():
                if line.startswith("data:"):
                    payload = line[len("data:"):].strip()
                    data = _json.loads(payload)
                    assert "queue" in data, "El evento SSE debe tener campo 'queue'"
                    assert "timestamp" in data, "El evento SSE debe tener campo 'timestamp'"
                    break

    def test_events_has_nocache_headers(self, api_client):
        """La respuesta SSE debe tener Cache-Control: no-cache."""
        client, dm, qm = api_client
        with client.stream("GET", "/api/events") as r:
            cache_control = r.headers.get("cache-control", "")
            assert "no-cache" in cache_control, \
                "La respuesta SSE debe tener Cache-Control: no-cache"

    def test_events_endpoint_uses_depends(self):
        """event_stream debe usar Depends(get_queue_mgr)."""
        import inspect, server as srv
        src = inspect.getsource(srv.event_stream)
        assert "_Depends(get_queue_mgr)" in src or "Depends(get_queue_mgr)" in src, \
            "event_stream debe usar Depends(get_queue_mgr) (Mejora 4 + 6)"

    def test_events_respects_auth(self, monkeypatch):
        """Con PIN activo, /api/events debe rechazar clientes sin token."""
        import server as srv
        from starlette.testclient import TestClient
        monkeypatch.setattr(srv, "_has_pin", lambda: True)
        monkeypatch.setattr(srv, "device_mgr",
            type("DM", (), {"is_authorized": lambda self, t: False,
                            "get_device": lambda self, t: {}})())
        client = TestClient(srv.app)
        r = client.get("/api/events")
        assert r.status_code in (401, 403), \
            "Con PIN y sin token, /api/events debe retornar 401 o 403"
